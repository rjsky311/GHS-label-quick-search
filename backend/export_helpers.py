from collections import Counter
import re
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter


EXPORT_BASE_HEADERS = [
    "CAS No.",
    "English Name",
    "Chinese Name",
    "GHS Pictograms",
    "Signal Word",
    "Hazard Statements",
    "Precautionary Statements",
]

EXPORT_TRUST_HEADERS = [
    "Data State",
    "Printable",
    "Needs Review",
    "Review Reasons",
    "Review Signal Count",
    "Primary Review Action",
    "Primary Source",
    "Report Count",
    "Retrieved At",
    "Cache State",
    "Reference Links",
    "Source Conflict",
    "Missing Trusted Chinese Name",
    "Multiple GHS Status",
    "Classification Selection",
]

EXPORT_DATA_HEADERS = [
    *EXPORT_BASE_HEADERS,
    *EXPORT_TRUST_HEADERS,
]

EXPORT_TRIAGE_SHEET_NAMES = {
    "ready": "Ready Rows",
    "needs_review": "Needs Review",
    "unresolved": "Unresolved",
}

EXPORT_REVIEW_REASON_LABELS = {
    "upstream_error": "Upstream retry needed",
    "unresolved_search": "Unresolved lookup",
    "no_ghs_data": "No GHS data",
    "ghs_text_no_pictograms": "GHS pictogram gap",
    "source_conflict": "Source conflict",
    "multiple_classifications": "Multiple GHS classifications",
    "missing_chinese_name": "Missing trusted Chinese name",
}

EXPORT_REVIEW_ACTION_LABELS = {
    "upstream_error": "Retry upstream source",
    "unresolved_search": "Send lookup gap to correction",
    "no_ghs_data": "Check SDS before relying on no-GHS result",
    "ghs_text_no_pictograms": "Review pictogram evidence",
    "source_conflict": "Inspect source conflict evidence",
    "multiple_classifications": "Confirm primary GHS classification",
    "missing_chinese_name": "Submit trusted Chinese-name evidence",
}


_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")
_CJK_RE = re.compile(r"[\u3400-\u4DBF\u4E00-\u9FFF\uF900-\uFAFF]")


def spreadsheet_safe(value: Any) -> str:
    """Neutralize spreadsheet formula injection in exported cells."""
    if value is None:
        return ""
    text = str(value)
    if text and text[0] in _FORMULA_TRIGGER_CHARS:
        return "'" + text
    return text


def _has_cjk_text(value: Any) -> bool:
    return bool(_CJK_RE.search(str(value or "")))


def _trusted_export_chinese_name(result: Dict[str, Any]) -> str:
    """Return a Chinese display name only when the payload contains CJK text."""
    for key in ("name_zh", "name_zh_tw"):
        value = str(result.get(key) or "").strip()
        if value and _has_cjk_text(value):
            return value
    return ""


def _has_export_ghs_data(result: Dict[str, Any]) -> bool:
    return bool(
        result.get("ghs_pictograms")
        or result.get("pictograms")
        or result.get("hazard_statements")
        or result.get("precautionary_statements")
        or result.get("signal_word")
    )


def _export_data_state(result: Dict[str, Any]) -> str:
    if result.get("upstream_error"):
        return "Upstream transient failure"
    if result.get("found") is False:
        return "Not found"
    if _has_export_ghs_data(result) and not (
        result.get("ghs_pictograms") or result.get("pictograms")
    ):
        return "Found with GHS text but no pictogram"
    if not _has_export_ghs_data(result):
        return "Found with no GHS classification data"
    return "Found with renderable GHS pictograms"


def _has_multiple_classifications(result: Dict[str, Any]) -> bool:
    return bool(
        result.get("has_multiple_classifications")
        or result.get("other_classifications")
    )


def _has_manual_classification_selection(result: Dict[str, Any]) -> bool:
    return result.get("selected_classification_index") is not None or bool(
        result.get("customNote")
    )


def _export_review_issue_types(result: Dict[str, Any]) -> List[str]:
    issue_types: List[str] = []
    if result.get("upstream_error"):
        issue_types.append("upstream_error")
    if result.get("found") is False:
        if not result.get("upstream_error"):
            issue_types.append("unresolved_search")
        return issue_types

    if not _has_export_ghs_data(result):
        issue_types.append("no_ghs_data")
    elif not (result.get("ghs_pictograms") or result.get("pictograms")):
        issue_types.append("ghs_text_no_pictograms")

    if result.get("source_conflict") or result.get("source_conflicts"):
        issue_types.append("source_conflict")
    if _has_multiple_classifications(result) and not _has_manual_classification_selection(result):
        issue_types.append("multiple_classifications")
    if result.get("name_en") and not _trusted_export_chinese_name(result):
        issue_types.append("missing_chinese_name")
    return issue_types


def _export_review_reasons(result: Dict[str, Any]) -> List[str]:
    return [
        EXPORT_REVIEW_REASON_LABELS[issue_type]
        for issue_type in _export_review_issue_types(result)
    ]


def _export_primary_review_action(issue_types: List[str]) -> str:
    if not issue_types:
        return "No review action"
    return EXPORT_REVIEW_ACTION_LABELS[issue_types[0]]


def _export_multiple_ghs_status(result: Dict[str, Any]) -> str:
    if not _has_multiple_classifications(result):
        return "Single or no alternate classification"
    if _has_manual_classification_selection(result):
        return "User-selected primary classification"
    return "Multiple classifications; using system-suggested primary"


def _export_trust_cells(result: Dict[str, Any]) -> List[str]:
    reference_links = result.get("reference_links")
    reference_count = (
        len(reference_links) if isinstance(reference_links, list) else 0
    )
    review_issue_types = _export_review_issue_types(result)
    review_reasons = [
        EXPORT_REVIEW_REASON_LABELS[issue_type]
        for issue_type in review_issue_types
    ]
    return [
        _export_data_state(result),
        "Yes" if result.get("found") is not False and _has_export_ghs_data(result) else "No",
        "Yes" if review_reasons else "No",
        "; ".join(review_reasons) if review_reasons else "No review reasons",
        str(len(review_issue_types)),
        _export_primary_review_action(review_issue_types),
        result.get("primary_source") or "Not recorded",
        result.get("primary_report_count") or "-",
        result.get("retrieved_at") or "Not recorded",
        "Cached" if result.get("cache_hit") else "Fresh or not recorded",
        f"{reference_count} reference link(s)"
        if reference_count
        else "No reference links",
        "Yes" if result.get("source_conflict") or result.get("source_conflicts") else "No",
        "Yes" if result.get("name_en") and not _trusted_export_chinese_name(result) else "No",
        _export_multiple_ghs_status(result),
        f"User-selected classification: {result.get('customNote')}"
        if result.get("customNote")
        else "Default primary classification",
    ]


def _statement_text(statement: Dict[str, Any]) -> str:
    return (
        statement.get("text_zh")
        or statement.get("text_en")
        or statement.get("text")
        or ""
    )


def _format_pictograms(result: Dict[str, Any]) -> str:
    pictograms = result.get("ghs_pictograms") or result.get("pictograms") or []
    if not pictograms:
        return "None"

    def pictogram_label(pictogram: Dict[str, Any]) -> str:
        name = (
            pictogram.get("name_zh")
            or pictogram.get("name_en")
            or pictogram.get("name")
            or ""
        )
        return f"{pictogram.get('code', '')} ({name})"

    return ", ".join(
        pictogram_label(pictogram) for pictogram in pictograms
    )


def _format_statements(
    statements: List[Dict[str, Any]] | None,
    *,
    empty_value: str,
    separator: str,
) -> str:
    if not statements:
        return empty_value
    return separator.join(
        f"{statement.get('code', '')}: {_statement_text(statement)}"
        for statement in statements
    )


def build_export_data_row(result: Dict[str, Any], *, multiline: bool = False) -> List[str]:
    separator = "\n" if multiline else "; "
    return [
        result.get("cas_number", ""),
        result.get("name_en", ""),
        _trusted_export_chinese_name(result),
        _format_pictograms(result),
        result.get("signal_word_zh") or result.get("signal_word") or "-",
        _format_statements(
            result.get("hazard_statements"),
            empty_value="No hazard statements",
            separator=separator,
        ),
        _format_statements(
            result.get("precautionary_statements"),
            empty_value="No precautionary statements",
            separator=separator,
        ),
        *_export_trust_cells(result),
    ]


def _is_export_ready_without_review(result: Dict[str, Any]) -> bool:
    return (
        result.get("found") is not False
        and _has_export_ghs_data(result)
        and not _export_review_reasons(result)
    )


def build_export_triage_sheets(
    results: List[Dict[str, Any]],
) -> dict[str, List[Dict[str, Any]]]:
    """Partition export rows into lab-manager triage sheets without overlap."""
    return {
        EXPORT_TRIAGE_SHEET_NAMES["ready"]: [
            result for result in results if _is_export_ready_without_review(result)
        ],
        EXPORT_TRIAGE_SHEET_NAMES["needs_review"]: [
            result
            for result in results
            if (
                result.get("upstream_error")
                or (
                    result.get("found") is not False
                    and _export_review_reasons(result)
                )
            )
        ],
        EXPORT_TRIAGE_SHEET_NAMES["unresolved"]: [
            result
            for result in results
            if result.get("found") is False and not result.get("upstream_error")
        ],
    }


def _apply_export_column_widths(worksheet) -> None:
    widths = [
        15,
        30,
        20,
        35,
        14,
        55,
        58,
        34,
        16,
        34,
        14,
        16,
        44,
        36,
        18,
        24,
        18,
        24,
        18,
        28,
        38,
        36,
    ]
    for index, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _style_results_header(worksheet, thin_border: Border) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid",
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, header in enumerate(EXPORT_DATA_HEADERS, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def add_export_results_sheet(
    workbook: Workbook,
    title: str,
    results: List[Dict[str, Any]],
    thin_border: Border,
    *,
    worksheet=None,
) -> None:
    sheet = worksheet or workbook.create_sheet(title)
    sheet.title = title
    _style_results_header(sheet, thin_border)
    sheet.freeze_panes = "A2"

    for row_index, result in enumerate(results, 2):
        for col_index, value in enumerate(
            build_export_data_row(result, multiline=True),
            1,
        ):
            cell = sheet.cell(
                row=row_index,
                column=col_index,
                value=spreadsheet_safe(value),
            )
            cell.border = thin_border
            if col_index >= 4:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    if results:
        sheet.auto_filter.ref = sheet.dimensions
    _apply_export_column_widths(sheet)


def _build_export_pilot_summary(
    results: List[Dict[str, Any]],
    *,
    export_scope: str = "visible",
    export_scope_label: str = "Visible filtered",
    export_count: int | None = None,
    source_total_count: int | None = None,
    visible_count: int | None = None,
) -> List[tuple[str, Any, str]]:
    review_reason_counts: Counter[str] = Counter()
    printable_count = 0
    review_signal_count = 0
    review_overlap_count = 0
    for result in results:
        if result.get("found") is not False and _has_export_ghs_data(result):
            printable_count += 1
        review_reasons = _export_review_reasons(result)
        review_reason_counts.update(review_reasons)
        review_signal_count += len(review_reasons)
        if len(review_reasons) > 1:
            review_overlap_count += 1

    needs_review_count = sum(
        1 for result in results if _export_review_reasons(result)
    )
    exported_count = len(results)
    total_rows = (
        max(int(source_total_count), exported_count)
        if source_total_count is not None
        else exported_count
    )
    visible_rows = (
        max(int(visible_count), 0)
        if visible_count is not None
        else exported_count
    )
    return [
        (
            "Export scope",
            export_scope_label or export_scope or "Visible filtered",
            "The user-selected scope for this workbook handoff.",
        ),
        (
            "Exported row count",
            exported_count,
            "Rows included in this exported file after scope filtering.",
        ),
        (
            "Total rows",
            total_rows,
            "Rows in the original batch/export context before scope filtering.",
        ),
        (
            "Visible filtered rows",
            visible_rows,
            "Rows visible in the table when the export preview was opened.",
        ),
        (
            "Printable rows",
            printable_count,
            "Rows with GHS content available for downstream label planning.",
        ),
        (
            "Needs review rows",
            needs_review_count,
            "Rows that should be checked before relying on the export.",
        ),
        (
            "Review signals",
            review_signal_count,
            "Total issue signals across review rows; this can exceed needs-review rows when one chemical has several issues.",
        ),
        (
            "Rows with multiple review signals",
            review_overlap_count,
            "Rows where more than one issue should be handled in the same triage pass.",
        ),
        (
            "Unresolved searches",
            review_reason_counts[EXPORT_REVIEW_REASON_LABELS["unresolved_search"]],
            "Queries that need dictionary curation or corrected input.",
        ),
        (
            "Missing trusted Chinese names",
            review_reason_counts[EXPORT_REVIEW_REASON_LABELS["missing_chinese_name"]],
            "Rows where Chinese name display is intentionally blank until reviewed.",
        ),
        (
            "Multiple GHS classifications",
            review_reason_counts[EXPORT_REVIEW_REASON_LABELS["multiple_classifications"]],
            "Rows where the user or maintainer should confirm the primary classification.",
        ),
        (
            "Source conflicts",
            review_reason_counts[EXPORT_REVIEW_REASON_LABELS["source_conflict"]],
            "Rows where PubChem/ECHA/manual evidence should remain visible.",
        ),
        (
            "No GHS data",
            review_reason_counts[EXPORT_REVIEW_REASON_LABELS["no_ghs_data"]],
            "Found records without GHS hazard data in the current result.",
        ),
        (
            "Text-only GHS without pictograms",
            review_reason_counts[EXPORT_REVIEW_REASON_LABELS["ghs_text_no_pictograms"]],
            "Rows with hazard text but no renderable pictogram in the result.",
        ),
        (
            "Upstream transient failures",
            review_reason_counts[EXPORT_REVIEW_REASON_LABELS["upstream_error"]],
            "Rows that should be retried before drawing a safety conclusion.",
        ),
    ]


def _add_export_pilot_summary_sheet(
    workbook: Workbook,
    results: List[Dict[str, Any]],
    thin_border: Border,
    *,
    export_scope: str = "visible",
    export_scope_label: str = "Visible filtered",
    export_count: int | None = None,
    source_total_count: int | None = None,
    visible_count: int | None = None,
) -> None:
    summary_ws = workbook.create_sheet("Pilot Summary")
    headers = ["Metric", "Value", "Pilot use"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, header in enumerate(headers, 1):
        cell = summary_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row, (metric, value, note) in enumerate(
        _build_export_pilot_summary(
            results,
            export_scope=export_scope,
            export_scope_label=export_scope_label,
            export_count=export_count,
            source_total_count=source_total_count,
            visible_count=visible_count,
        ),
        2,
    ):
        summary_ws.cell(row=row, column=1, value=spreadsheet_safe(metric)).border = thin_border
        safe_value = spreadsheet_safe(value) if isinstance(value, str) else value
        summary_ws.cell(row=row, column=2, value=safe_value).border = thin_border
        note_cell = summary_ws.cell(row=row, column=3, value=spreadsheet_safe(note))
        note_cell.border = thin_border
        note_cell.alignment = Alignment(wrap_text=True)

    summary_ws.column_dimensions["A"].width = 34
    summary_ws.column_dimensions["B"].width = 14
    summary_ws.column_dimensions["C"].width = 72
