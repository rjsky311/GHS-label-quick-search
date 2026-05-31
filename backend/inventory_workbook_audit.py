from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import load_workbook

from api_validation import CAS_FORMAT_PATTERN, has_valid_cas_checksum
from candidate_discovery import build_candidate_bundle, clean_text, has_cjk
from chemical_dict import CAS_TO_EN, CAS_TO_ZH

AUDIT_SOURCE = "inventory-workbook-audit"
HANDOFF_RECORD_KEYS = (
    "invalidCas",
    "workbookChineseNameCandidates",
    "unknownSeedDictionary",
    "missingSeedChineseName",
    "casCleanupSignals",
    "duplicates",
)

CAS_DIGITS_PATTERN = re.compile(r"^\d{5,10}$")
CAS_PREFIX_PATTERN = re.compile(r"^cas\s*(?:no\.?|number|#|[:\uff1a])?\s*", re.I)
DASH_PATTERN = re.compile(r"[\u2010-\u2015\u2212\ufe58\ufe63\uff0d]")
TRAILING_PUNCTUATION_PATTERN = re.compile(r"[\.,;:\uff0c\u3002\uff1b\uff1a]+$")
HEADER_STRIP_PATTERN = re.compile(r"[\s\.:#_/\-()\uff08\uff09]+")

KNOWN_CAS_HEADER_CELLS = frozenset(
    {
        "cas",
        "casno",
        "casnumber",
        "cas\u7de8\u865f",
        "cas\u7f16\u53f7",
        "cas\u865f",
        "cas\u53f7",
        "cas\u865f\u78bc",
        "cas\u53f7\u7801",
        "cas\u767b\u9304\u865f",
        "cas\u767b\u5f55\u53f7",
        "cas\u767b\u8a18\u865f",
        "cas\u767b\u8bb0\u53f7",
    }
)

ENGLISH_NAME_HEADER_CELLS = frozenset(
    {
        "name",
        "englishname",
        "nameen",
        "\u82f1\u6587",
        "\u82f1\u6587\u540d",
        "\u82f1\u6587\u540d\u7a31",
        "\u82f1\u6587\u540d\u79f0",
        "\u82f1\u6587\u54c1\u540d",
        "\u85e5\u54c1\u540d\u7a31",
        "\u836f\u54c1\u540d\u79f0",
    }
)

CHINESE_NAME_HEADER_CELLS = frozenset(
    {
        "\u4e2d\u6587",
        "\u4e2d\u6587\u540d",
        "\u4e2d\u6587\u540d\u7a31",
        "\u4e2d\u6587\u540d\u79f0",
        "\u4e2d\u6587\u54c1\u540d",
        "\u4e2d\u540d",
    }
)


def to_half_width(value: Any = "") -> str:
    text = normalize_cell_text(value)
    return "".join(
        chr(ord(char) - 0xFEE0) if 0xFF01 <= ord(char) <= 0xFF5E else char
        for char in text
    )


def normalize_cell_text(value: Any = "") -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def normalize_header_cell(value: Any = "") -> str:
    return HEADER_STRIP_PATTERN.sub("", to_half_width(value).strip().lower())


def is_cas_header_cell(value: Any = "") -> bool:
    return normalize_header_cell(value) in KNOWN_CAS_HEADER_CELLS


def find_header_index(row: Iterable[Any], candidates: frozenset[str]) -> Optional[int]:
    for index, value in enumerate(row):
        if normalize_header_cell(value) in candidates:
            return index
    return None


def rehyphenate_cas_digits(value: str) -> str:
    digits = str(value or "").strip()
    if not CAS_DIGITS_PATTERN.fullmatch(digits):
        return ""
    first_group = digits[:-3]
    if len(first_group) < 2 or len(first_group) > 7:
        return ""
    return f"{first_group}-{digits[-3:-1]}-{digits[-1]}"


def canonicalize_cas_leading_zeros(value: str) -> str:
    if not CAS_FORMAT_PATTERN.fullmatch(value):
        return ""
    first_group, middle_group, check_digit = value.split("-")
    canonical_first_group = first_group.lstrip("0")
    if not canonical_first_group or canonical_first_group == first_group:
        return ""
    return f"{canonical_first_group}-{middle_group}-{check_digit}"


def normalize_cas_token_detailed(value: Any = "") -> dict[str, Any]:
    raw = normalize_cell_text(value)
    raw_normalized = (
        DASH_PATTERN.sub("-", to_half_width(raw))
        .strip()
        .replace("\u3000", "")
    )
    raw_normalized = CAS_PREFIX_PATTERN.sub("", raw_normalized)
    raw_normalized = re.sub(r"\s+", "", raw_normalized)
    raw_normalized = re.sub(r"\.0+$", "", raw_normalized)
    raw_normalized = TRAILING_PUNCTUATION_PATTERN.sub("", raw_normalized).strip()

    rehyphenated = rehyphenate_cas_digits(raw_normalized)
    normalized_candidate = rehyphenated or raw_normalized
    leading_zero_canonical = canonicalize_cas_leading_zeros(normalized_candidate)
    normalized = leading_zero_canonical or normalized_candidate
    format_valid = bool(CAS_FORMAT_PATTERN.fullmatch(normalized))
    checksum_valid = bool(format_valid and has_valid_cas_checksum(normalized))
    reason = ""
    if not format_valid:
        reason = "format"
    elif not checksum_valid:
        reason = "checksum"

    return {
        "raw": raw,
        "rawNormalized": raw_normalized,
        "normalized": normalized,
        "wasRehyphenated": bool(rehyphenated and rehyphenated != raw_normalized),
        "wasLeadingZeroCanonicalized": bool(leading_zero_canonical),
        "formatValid": format_valid,
        "checksumValid": checksum_valid,
        "valid": checksum_valid,
        "reason": reason,
    }


def _row_value(row: tuple[Any, ...], index: Optional[int]) -> str:
    if index is None or index >= len(row):
        return ""
    return clean_text(normalize_cell_text(row[index]), max_length=240)


def _append_example(
    examples: dict[str, list[dict[str, Any]]],
    key: str,
    item: dict[str, Any],
    *,
    max_examples: int,
) -> None:
    if len(examples[key]) < max_examples:
        examples[key].append(item)


def _append_handoff_record(
    records: Optional[dict[str, list[dict[str, Any]]]],
    key: str,
    item: dict[str, Any],
) -> None:
    if records is None:
        return
    records[key].append(item)


def _detect_sheet_mapping(
    worksheet: Any,
    *,
    max_header_scan_rows: int,
) -> Optional[dict[str, Optional[int]]]:
    for row_index, row in enumerate(
        worksheet.iter_rows(max_row=max_header_scan_rows, values_only=True),
        start=1,
    ):
        cas_index = next(
            (index for index, value in enumerate(row) if is_cas_header_cell(value)),
            None,
        )
        if cas_index is None:
            continue
        return {
            "headerRow": row_index,
            "casIndex": cas_index,
            "nameEnIndex": find_header_index(row, ENGLISH_NAME_HEADER_CELLS),
            "nameZhIndex": find_header_index(row, CHINESE_NAME_HEADER_CELLS),
        }
    return None


def _candidate_from_workbook_row(
    *,
    cas_number: str,
    name_en: str,
    name_zh: str,
    sheet_name: str,
    row_number: int,
) -> dict[str, Any]:
    candidate = build_candidate_bundle(
        candidate_type="missing-chinese-name",
        cas_number=cas_number,
        name_en=name_en,
        name_zh=name_zh,
        evidence_type="Inventory workbook Chinese name",
        review_notes=(
            "Inventory workbook candidate only. Verify against SDS, supplier label, "
            "or authoritative local source before approving a manual dictionary entry."
        ),
        source=AUDIT_SOURCE,
    )
    candidate["sheet"] = sheet_name
    candidate["row"] = row_number
    return candidate


def _inventory_action(
    *,
    key: str,
    count: int,
    severity: str,
    target_example_key: str | list[str],
    title: str,
    next_action: str,
    blocks_batch_use: bool = False,
) -> Optional[dict[str, Any]]:
    if count <= 0:
        return None
    target_example_keys = (
        target_example_key if isinstance(target_example_key, list) else [target_example_key]
    )
    return {
        "key": key,
        "count": count,
        "severity": severity,
        "targetExampleKey": target_example_keys[0] if target_example_keys else "",
        "targetExampleKeys": target_example_keys,
        "title": title,
        "nextAction": next_action,
        "blocksBatchUse": blocks_batch_use,
        "review_required": True,
        "public_data_changed": False,
    }


def build_inventory_action_queue(summary: dict[str, Any]) -> list[dict[str, Any]]:
    """Return maintainer actions derived from the audit summary.

    The queue is intentionally conservative: it describes work to review, fix,
    or use as QA evidence, but never implies that workbook values are approved
    public dictionary data.
    """

    missing_without_candidate = max(
        int(summary.get("missingSeedChineseNameRows") or 0)
        - int(summary.get("workbookChineseNameCandidateRows") or 0),
        0,
    )
    actions = [
        _inventory_action(
            key="fix-invalid-cas",
            count=int(summary.get("invalidCasCount") or 0),
            severity="blocking",
            target_example_key="invalidCas",
            title="Fix invalid CAS cells before batch use",
            next_action=(
                "Correct format/checksum problems in the source workbook; "
                "invalid CAS rows cannot be searched, printed, exported, or "
                "turned into dictionary candidates."
            ),
            blocks_batch_use=True,
        ),
        _inventory_action(
            key="review-workbook-chinese-candidates",
            count=int(summary.get("workbookChineseNameCandidateRows") or 0),
            severity="review",
            target_example_key="workbookChineseNameCandidates",
            title="Review workbook Chinese-name candidates",
            next_action=(
                "Verify each workbook-provided Chinese name against SDS, "
                "supplier label, or regulatory evidence before converting it "
                "into a pending or approved manual dictionary entry."
            ),
        ),
        _inventory_action(
            key="triage-unknown-seed-dictionary",
            count=int(summary.get("unknownSeedDictionaryRows") or 0),
            severity="review",
            target_example_key="unknownSeedDictionary",
            title="Triage CAS rows outside the seed dictionary",
            next_action=(
                "Decide whether each CAS/name pair needs a correction request, "
                "manual-entry candidate, alias, or no action. Do not import "
                "workbook identity fields directly into public lookup."
            ),
        ),
        _inventory_action(
            key="collect-missing-chinese-name-evidence",
            count=missing_without_candidate,
            severity="evidence",
            target_example_key="missingSeedChineseName",
            title="Collect evidence for missing Chinese names",
            next_action=(
                "For rows without a usable workbook Chinese-name candidate, "
                "collect SDS, supplier, regulatory, or reviewed local evidence "
                "before creating manual dictionary entries."
            ),
        ),
        _inventory_action(
            key="confirm-cas-cleanup-coverage",
            count=int(summary.get("casCleanupSignalRows") or 0),
            severity="qa",
            target_example_key=["rehyphenatedCas", "leadingZeroCas"],
            title="Keep CAS cleanup covered by parser and production QA",
            next_action=(
                "Use numeric-CAS and leading-zero examples as parser/production "
                "QA fixtures so spreadsheet artifacts become canonical CAS "
                "before search, review, print, and export."
            ),
        ),
        _inventory_action(
            key="deduplicate-workbook-rows",
            count=int(summary.get("duplicateValidCasRows") or 0),
            severity="cleanup",
            target_example_key="duplicates",
            title="Deduplicate repeated valid CAS rows",
            next_action=(
                "Deduplicate before lab-manager handoff when the repeated rows "
                "do not represent separate physical containers or locations."
            ),
        ),
    ]
    return [action for action in actions if action]


def audit_inventory_workbook(
    workbook_path: str | Path,
    *,
    max_examples: int = 20,
    max_header_scan_rows: int = 40,
    include_handoff_records: bool = False,
) -> dict[str, Any]:
    path = Path(workbook_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    seen_valid_cas: set[str] = set()
    examples: dict[str, list[dict[str, Any]]] = {
        "invalidCas": [],
        "rehyphenatedCas": [],
        "leadingZeroCas": [],
        "duplicates": [],
        "unknownSeedDictionary": [],
        "missingSeedChineseName": [],
        "workbookChineseNameCandidates": [],
    }
    handoff_records = (
        {key: [] for key in HANDOFF_RECORD_KEYS} if include_handoff_records else None
    )
    sheet_summaries: list[dict[str, Any]] = []
    summary = {
        "sheetCount": len(workbook.sheetnames),
        "sheetsWithCasColumn": 0,
        "skippedSheets": 0,
        "casCellCount": 0,
        "validCasRowCount": 0,
        "uniqueValidCasCount": 0,
        "duplicateValidCasRows": 0,
        "invalidCasCount": 0,
        "casCleanupSignalRows": 0,
        "rehyphenatedCasCount": 0,
        "leadingZeroCasCount": 0,
        "knownSeedDictionaryRows": 0,
        "unknownSeedDictionaryRows": 0,
        "missingSeedChineseNameRows": 0,
        "workbookChineseNameCandidateRows": 0,
    }

    try:
        for worksheet in workbook.worksheets:
            mapping = _detect_sheet_mapping(
                worksheet,
                max_header_scan_rows=max_header_scan_rows,
            )
            if not mapping:
                summary["skippedSheets"] += 1
                sheet_summaries.append(
                    {
                        "sheet": worksheet.title,
                        "status": "skipped_no_cas_header",
                        "casCellCount": 0,
                        "validCasRowCount": 0,
                        "invalidCasCount": 0,
                    }
                )
                continue

            summary["sheetsWithCasColumn"] += 1
            sheet_summary = {
                "sheet": worksheet.title,
                "status": "audited",
                "headerRow": mapping["headerRow"],
                "casColumnIndex": mapping["casIndex"],
                "nameEnColumnIndex": mapping["nameEnIndex"],
                "nameZhColumnIndex": mapping["nameZhIndex"],
                "casCellCount": 0,
                "validCasRowCount": 0,
                "invalidCasCount": 0,
                "duplicateValidCasRows": 0,
                "casCleanupSignalRows": 0,
                "rehyphenatedCasCount": 0,
                "leadingZeroCasCount": 0,
                "workbookChineseNameCandidateRows": 0,
            }

            for row_number, row in enumerate(
                worksheet.iter_rows(
                    min_row=int(mapping["headerRow"] or 1) + 1,
                    values_only=True,
                ),
                start=int(mapping["headerRow"] or 1) + 1,
            ):
                cas_index = int(mapping["casIndex"] or 0)
                raw_cas = row[cas_index] if cas_index < len(row) else ""
                if not clean_text(normalize_cell_text(raw_cas), max_length=80):
                    continue

                detail = normalize_cas_token_detailed(raw_cas)
                row_context = {
                    "sheet": worksheet.title,
                    "row": row_number,
                    "raw": detail["raw"],
                    "normalized": detail["normalized"],
                }
                summary["casCellCount"] += 1
                sheet_summary["casCellCount"] += 1

                if not detail["valid"]:
                    summary["invalidCasCount"] += 1
                    sheet_summary["invalidCasCount"] += 1
                    invalid_item = {**row_context, "reason": detail["reason"]}
                    _append_example(
                        examples,
                        "invalidCas",
                        invalid_item,
                        max_examples=max_examples,
                    )
                    _append_handoff_record(handoff_records, "invalidCas", invalid_item)
                    continue

                cas_number = str(detail["normalized"])
                name_en = _row_value(row, mapping["nameEnIndex"]) or CAS_TO_EN.get(
                    cas_number, ""
                )
                workbook_name_zh = _row_value(row, mapping["nameZhIndex"])
                seed_name_zh = CAS_TO_ZH.get(cas_number, "")
                known_in_seed = bool(CAS_TO_EN.get(cas_number) or seed_name_zh)

                summary["validCasRowCount"] += 1
                sheet_summary["validCasRowCount"] += 1
                if known_in_seed:
                    summary["knownSeedDictionaryRows"] += 1
                else:
                    summary["unknownSeedDictionaryRows"] += 1
                    unknown_item = {
                        **row_context,
                        "nameEn": name_en,
                        "nameZh": workbook_name_zh,
                    }
                    _append_example(
                        examples,
                        "unknownSeedDictionary",
                        unknown_item,
                        max_examples=max_examples,
                    )
                    _append_handoff_record(
                        handoff_records,
                        "unknownSeedDictionary",
                        unknown_item,
                    )

                if detail["wasRehyphenated"]:
                    summary["rehyphenatedCasCount"] += 1
                    sheet_summary["rehyphenatedCasCount"] += 1
                    _append_example(
                        examples,
                        "rehyphenatedCas",
                        {
                            **row_context,
                            "rawNormalized": detail["rawNormalized"],
                        },
                        max_examples=max_examples,
                    )

                if detail["wasLeadingZeroCanonicalized"]:
                    summary["leadingZeroCasCount"] += 1
                    sheet_summary["leadingZeroCasCount"] += 1
                    _append_example(
                        examples,
                        "leadingZeroCas",
                        {
                            **row_context,
                            "rawNormalized": detail["rawNormalized"],
                            "normalized": cas_number,
                        },
                        max_examples=max_examples,
                    )

                if cas_number in seen_valid_cas:
                    summary["duplicateValidCasRows"] += 1
                    sheet_summary["duplicateValidCasRows"] += 1
                    duplicate_item = {**row_context, "reason": "duplicate-valid-cas"}
                    _append_example(
                        examples,
                        "duplicates",
                        duplicate_item,
                        max_examples=max_examples,
                    )
                    _append_handoff_record(handoff_records, "duplicates", duplicate_item)
                else:
                    seen_valid_cas.add(cas_number)
                    summary["uniqueValidCasCount"] += 1

                if detail["wasRehyphenated"] or detail["wasLeadingZeroCanonicalized"]:
                    summary["casCleanupSignalRows"] += 1
                    sheet_summary["casCleanupSignalRows"] += 1
                    _append_handoff_record(
                        handoff_records,
                        "casCleanupSignals",
                        {
                            **row_context,
                            "rawNormalized": detail["rawNormalized"],
                            "wasRehyphenated": detail["wasRehyphenated"],
                            "wasLeadingZeroCanonicalized": detail[
                                "wasLeadingZeroCanonicalized"
                            ],
                        },
                    )

                if not has_cjk(seed_name_zh):
                    summary["missingSeedChineseNameRows"] += 1
                    missing_item = {
                        **row_context,
                        "nameEn": name_en,
                        "workbookNameZh": workbook_name_zh,
                    }
                    _append_example(
                        examples,
                        "missingSeedChineseName",
                        missing_item,
                        max_examples=max_examples,
                    )
                    _append_handoff_record(
                        handoff_records,
                        "missingSeedChineseName",
                        missing_item,
                    )
                    if has_cjk(workbook_name_zh):
                        summary["workbookChineseNameCandidateRows"] += 1
                        sheet_summary["workbookChineseNameCandidateRows"] += 1
                        candidate = _candidate_from_workbook_row(
                            cas_number=cas_number,
                            name_en=name_en,
                            name_zh=workbook_name_zh,
                            sheet_name=worksheet.title,
                            row_number=row_number,
                        )
                        _append_example(
                            examples,
                            "workbookChineseNameCandidates",
                            candidate,
                            max_examples=max_examples,
                        )
                        _append_handoff_record(
                            handoff_records,
                            "workbookChineseNameCandidates",
                            candidate,
                        )

            sheet_summaries.append(sheet_summary)
    finally:
        workbook.close()

    payload = {
        "schemaVersion": 1,
        "dryRun": True,
        "publicDataChanged": False,
        "source": AUDIT_SOURCE,
        "workbook": str(path),
        "summary": summary,
        "actionQueue": build_inventory_action_queue(summary),
        "sheetSummaries": sheet_summaries,
        "examples": examples,
    }
    if handoff_records is not None:
        payload["handoffRecords"] = handoff_records
    return payload
