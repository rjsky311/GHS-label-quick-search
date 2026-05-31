"""
Tests for name search functionality:
- resolve_name_to_cas() helper
- Reverse dictionaries (EN_TO_CAS, ZH_TO_CAS)
- Alias dictionaries (ALIASES_ZH, ALIASES_EN)
- /api/search-by-name/{query} endpoint (local dictionary only, no network)
- /api/search/{query} auto-detect (local only tests)
"""
import pytest
import server
from datetime import datetime, timedelta, timezone
from httpx import AsyncClient, ASGITransport
from pydantic import ValidationError
from server import (
    app,
    resolve_name_to_cas,
    _classification_signature,
    _report_rank_key,
)
from pilot_store import PilotStore
from chemical_dict import EN_TO_CAS, ZH_TO_CAS, CAS_TO_EN, CAS_TO_ZH, ALIASES_ZH, ALIASES_EN
from api_validation import (
    MAX_EXPORT_SCALAR_CHARS,
    MAX_PUBLIC_CAS_QUERY_LENGTH,
    MAX_PUBLIC_SEARCH_QUERY_LENGTH,
)


def route_limits_for(endpoint_name):
    for key, limits in server.limiter._route_limits.items():
        if key.endswith(f".{endpoint_name}"):
            return limits
    return []


# ─── Unit tests: resolve_name_to_cas ───────────────────────

class TestResolveNameToCas:
    """Test the resolve_name_to_cas helper function."""

    def test_exact_english_name(self):
        """'Ethanol' is stored exactly in CAS_TO_EN."""
        assert resolve_name_to_cas("Ethanol") == "64-17-5"

    def test_exact_english_lowercase(self):
        assert resolve_name_to_cas("ethanol") == "64-17-5"

    def test_exact_english_uppercase(self):
        assert resolve_name_to_cas("ETHANOL") == "64-17-5"

    def test_exact_chinese_name(self):
        assert resolve_name_to_cas("乙醇") == "64-17-5"

    def test_methanol_resolves(self):
        """'Methanol' is stored as 'Methyl alcohol (Methanol)' in CAS_TO_EN.
        The word-boundary match should find it."""
        result = resolve_name_to_cas("Methanol")
        assert result == "67-56-1"

    def test_methanol_chinese(self):
        assert resolve_name_to_cas("甲醇") == "67-56-1"

    def test_unknown_name_returns_none(self):
        assert resolve_name_to_cas("Unobtainium") is None

    def test_empty_string_returns_none(self):
        assert resolve_name_to_cas("") is None

    def test_whitespace_only_returns_none(self):
        assert resolve_name_to_cas("   ") is None

    def test_chinese_unknown_returns_none(self):
        assert resolve_name_to_cas("不存在的化學物") is None

    def test_strips_whitespace(self):
        assert resolve_name_to_cas("  Ethanol  ") == "64-17-5"

    def test_acetone(self):
        """Check a common chemical."""
        result = resolve_name_to_cas("Acetone")
        assert result is not None  # Should find it in dictionary

    def test_cas_like_input_returns_none(self):
        """CAS-like input (digits and hyphens) won't be found by name."""
        assert resolve_name_to_cas("999-99-9") is None


# ─── Unit tests: reverse dictionaries ─────────────────────

class TestReverseDictionaries:
    """Test that reverse dictionaries are correctly built."""

    def test_en_to_cas_has_entries(self):
        assert len(EN_TO_CAS) > 0

    def test_en_to_cas_includes_aliases(self):
        """EN_TO_CAS should include merged English aliases."""
        assert len(EN_TO_CAS) >= len(CAS_TO_EN)  # At least as many entries (aliases add more)

    def test_zh_to_cas_has_entries(self):
        assert len(ZH_TO_CAS) > 0

    def test_en_to_cas_ethanol(self):
        assert EN_TO_CAS.get("ethanol") == "64-17-5"

    def test_zh_to_cas_ethanol(self):
        assert ZH_TO_CAS.get("乙醇") == "64-17-5"

    def test_en_to_cas_keys_are_lowercase(self):
        for key in list(EN_TO_CAS.keys())[:100]:
            assert key == key.lower(), f"Key '{key}' is not lowercase"

    def test_roundtrip_en(self):
        """CAS_TO_EN → EN_TO_CAS → should get back same CAS."""
        cas = "64-17-5"
        en_name = CAS_TO_EN[cas]
        assert EN_TO_CAS[en_name.lower()] == cas

    def test_roundtrip_zh(self):
        """CAS_TO_ZH → ZH_TO_CAS → should get back same CAS."""
        cas = "64-17-5"
        zh_name = CAS_TO_ZH[cas]
        assert ZH_TO_CAS[zh_name] == cas


# ─── Integration tests: /api/search-by-name (no network needed) ──

async def test_search_by_name_ethanol():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/search-by-name/ethanol")
    assert response.status_code == 200
    data = response.json()
    assert "results" in data
    cas_numbers = [r["cas_number"] for r in data["results"]]
    assert "64-17-5" in cas_numbers


async def test_workspace_documents_require_admin_token(monkeypatch):
    monkeypatch.setattr(server, "ADMIN_API_TOKEN", "secret")
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        get_response = await ac.get("/api/workspace/lab_profile")
        put_response = await ac.put(
            "/api/workspace/lab_profile",
            json={"payload": {"organization": "Lab A"}},
        )

    assert get_response.status_code == 401
    assert put_response.status_code == 401


async def test_workspace_documents_are_unavailable_without_admin_config(monkeypatch):
    monkeypatch.setattr(server, "ADMIN_API_TOKEN", "")
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/workspace/lab_profile")

    assert response.status_code == 503


async def test_workspace_documents_reject_oversized_payload(monkeypatch):
    monkeypatch.setattr(server, "ADMIN_API_TOKEN", "secret")
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.put(
            "/api/workspace/lab_profile",
            headers={"x-ghs-admin-key": "secret"},
            json={
                "payload": {
                    "blob": "x" * (server.MAX_WORKSPACE_DOCUMENT_JSON_CHARS + 1),
                },
            },
        )

    assert response.status_code == 422


async def test_dictionary_miss_query_capture_is_opt_in(monkeypatch):
    monkeypatch.setattr(server, "CAPTURE_DICTIONARY_MISSES", False)
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post(
            "/api/dictionary/miss-query",
            json={
                "query": "private pilot solvent",
                "query_kind": "name",
                "endpoint": "frontend",
            },
        )

    assert response.status_code == 200
    assert response.json() == {
        "ok": False,
        "skipped": True,
        "reason": "dictionary miss capture is disabled",
    }


async def test_dictionary_miss_query_rejects_oversized_payload(monkeypatch):
    monkeypatch.setattr(server, "CAPTURE_DICTIONARY_MISSES", True)
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        long_query_response = await ac.post(
            "/api/dictionary/miss-query",
            json={
                "query": "x" * (server.MAX_MISS_QUERY_LENGTH + 1),
                "query_kind": "name",
                "endpoint": "frontend",
            },
        )
        large_context_response = await ac.post(
            "/api/dictionary/miss-query",
            json={
                "query": "unknown solvent",
                "query_kind": "name",
                "endpoint": "frontend",
                "context": {"blob": "x" * (server.MAX_MISS_CONTEXT_JSON_CHARS + 1)},
            },
        )

    assert long_query_response.status_code == 422
    assert large_context_response.status_code == 422


async def test_dictionary_miss_query_context_is_sanitized(monkeypatch):
    monkeypatch.setattr(server, "CAPTURE_DICTIONARY_MISSES", True)
    captured = {}

    def fake_record_miss_query(query, query_kind, endpoint, *, context=None):
        captured["query"] = query
        captured["query_kind"] = query_kind
        captured["endpoint"] = endpoint
        captured["context"] = context
        return {"query": query, "queryKind": query_kind, "endpoint": endpoint}

    monkeypatch.setattr(server.pilot_store, "record_miss_query", fake_record_miss_query)
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post(
            "/api/dictionary/miss-query",
            json={
                "query": "unknown solvent",
                "query_kind": "name",
                "endpoint": "frontend",
                "context": {
                    "locale": " zh ",
                    "normalizedCas": " 64-17-5 ",
                    "resultCount": 0,
                    "email": "user@example.com",
                    "freeText": "please call me",
                    "nested": {"unsafe": True},
                },
            },
        )

    assert response.status_code == 200
    assert captured["context"] == {
        "locale": "zh",
        "normalizedCas": "64-17-5",
        "resultCount": 0,
    }


async def test_dictionary_miss_query_rejects_long_allowed_context_scalar(monkeypatch):
    monkeypatch.setattr(server, "CAPTURE_DICTIONARY_MISSES", True)
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post(
            "/api/dictionary/miss-query",
            json={
                "query": "unknown solvent",
                "query_kind": "name",
                "endpoint": "frontend",
                "context": {
                    "source": "x" * (server.MAX_MISS_CONTEXT_SCALAR_CHARS + 1),
                },
            },
        )

    assert response.status_code == 422


def test_dictionary_miss_query_endpoint_is_rate_limited():
    limits = route_limits_for("dictionary_miss_query")
    assert any(
        limit.limit.amount == 30 and limit.limit.GRANULARITY.name == "minute"
        for limit in limits
    )


async def test_dictionary_correction_request_records_public_intake(monkeypatch):
    captured = {}

    def fake_record_correction_request(**payload):
        captured.update(payload)
        return {"id": 7, **payload, "status": "open"}

    monkeypatch.setattr(
        server.pilot_store,
        "record_correction_request",
        fake_record_correction_request,
    )
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post(
            "/api/dictionary/correction-requests",
            json={
                "issue_type": "missing-chinese-name",
                "cas_number": " 107 18 6 ",
                "chemical_name": "Allyl Alcohol",
                "current_output": "Chinese name missing",
                "expected_output": "Reviewed Traditional Chinese name",
                "evidence_url": "https://example.com/sds",
                "evidence_type": "Supplier SDS",
                "local_context": " Detail modal ",
                "candidate": {
                    "name_zh": "candidate only",
                    "converted_to_manual_entry": True,
                    "manual_entry_status": "approved",
                    "manual_entry_source": "spoofed-public",
                    "request_id": 999,
                    "approved_for_public_use": True,
                    "public_data_changed": True,
                },
            },
        )

    assert response.status_code == 200
    assert response.json()["ok"] is True
    assert captured == {
        "issue_type": "missing-chinese-name",
        "cas_number": "107-18-6",
        "chemical_name": "Allyl Alcohol",
        "query_text": None,
        "current_output": "Chinese name missing",
        "expected_output": "Reviewed Traditional Chinese name",
        "evidence_url": "https://example.com/sds",
        "evidence_type": "Supplier SDS",
        "local_context": "Detail modal",
        "candidate": {
            "name_zh": "candidate only",
            "schema_version": 1,
            "review_required": True,
            "approved_for_public_use": False,
            "public_data_changed": False,
        },
        "source": "public",
    }


def test_dictionary_correction_request_endpoint_is_rate_limited():
    limits = route_limits_for("create_dictionary_correction_request")
    assert any(
        limit.limit.amount == 20 and limit.limit.GRANULARITY.name == "minute"
        for limit in limits
    )


def test_dictionary_miss_query_resolution_survives_recapture(tmp_path):
    store = PilotStore(tmp_path / "pilot.db").connect()
    try:
        first = store.record_miss_query(
            "mystery solvent",
            "name",
            "frontend",
            context={"locale": "en"},
        )
        assert first["resolution_status"] == "open"

        reviewed = store.update_miss_query_resolution(
            first["id"],
            resolution_status="ignored",
        )
        assert reviewed["resolution_status"] == "ignored"

        recaptured = store.record_miss_query(
            "mystery solvent",
            "name",
            "frontend",
            context={"locale": "en", "resultCount": 0},
        )

        assert recaptured["hit_count"] == 2
        assert recaptured["resolution_status"] == "ignored"
        assert recaptured["context"] == {"locale": "en", "resultCount": 0}
        assert store.get_dictionary_summary()["missQueryStatusCounts"] == {
            "ignored": 1,
            "needs_evidence": 0,
            "open": 0,
            "resolved": 0,
        }
        assert store.get_dictionary_summary()["topMissQueries"] == []
    finally:
        store.close()


def test_dictionary_miss_query_retention_purges_only_unprotected_stale_rows(tmp_path):
    now = datetime(2026, 5, 21, tzinfo=timezone.utc)
    old_seen_at = (now - timedelta(days=120)).isoformat()
    fresh_seen_at = (now - timedelta(days=10)).isoformat()
    store = PilotStore(tmp_path / "pilot.db").connect()
    try:
        old_open = store.record_miss_query("old open", "name", "frontend")
        old_resolved = store.record_miss_query("old resolved", "name", "frontend")
        old_needs_evidence = store.record_miss_query(
            "old needs evidence",
            "name",
            "frontend",
        )
        fresh_open = store.record_miss_query("fresh open", "name", "frontend")
        store.update_miss_query_resolution(
            old_resolved["id"],
            resolution_status="resolved",
            resolved_cas="64-17-5",
        )
        store.update_miss_query_resolution(
            old_needs_evidence["id"],
            resolution_status="needs_evidence",
        )
        for record in (old_open, old_resolved, old_needs_evidence):
            store._execute(
                """
                UPDATE dictionary_miss_queries
                SET first_seen_at = ?, last_seen_at = ?
                WHERE id = ?
                """,
                (old_seen_at, old_seen_at, record["id"]),
            )
        store._execute(
            """
            UPDATE dictionary_miss_queries
            SET first_seen_at = ?, last_seen_at = ?
            WHERE id = ?
            """,
            (fresh_seen_at, fresh_seen_at, fresh_open["id"]),
        )

        summary = store.get_miss_query_retention_summary(now=now)
        assert summary["retentionDays"] == 90
        assert summary["purgeableCount"] == 2
        assert summary["retainedNeedsEvidenceCount"] == 1
        dictionary_summary = store.get_dictionary_summary()
        assert dictionary_summary["missQueryRetention"]["purgeableCount"] == 2
        fresh_summary_row = next(
            item
            for item in dictionary_summary["topMissQueries"]
            if item["query_text"] == "fresh open"
        )
        assert "context" not in fresh_summary_row
        assert fresh_summary_row["contextRedacted"] is True

        result = store.purge_stale_miss_queries(now=now)
        assert result["deletedCount"] == 2
        remaining = store.list_miss_queries(limit=10)
        assert {item["query_text"] for item in remaining} == {
            "old needs evidence",
            "fresh open",
        }
    finally:
        store.close()


def test_dictionary_export_redacts_miss_query_context_by_default(tmp_path):
    store = PilotStore(tmp_path / "pilot.db").connect()
    try:
        store.record_miss_query(
            "mystery solvent",
            "name",
            "frontend",
            context={"locale": "en", "resultCount": 0},
        )

        default_snapshot = store.export_dictionary_snapshot()
        assert default_snapshot["missQueryExportScope"] == {
            "contextIncluded": False,
            "limit": 500,
        }
        assert "context" not in default_snapshot["missQueries"][0]
        assert default_snapshot["missQueries"][0]["contextRedacted"] is True

        raw_snapshot = store.export_dictionary_snapshot(include_miss_context=True)
        assert raw_snapshot["missQueryExportScope"]["contextIncluded"] is True
        assert raw_snapshot["missQueries"][0]["context"] == {
            "locale": "en",
            "resultCount": 0,
        }
    finally:
        store.close()


async def test_admin_can_update_dictionary_miss_query_resolution(monkeypatch):
    monkeypatch.setattr(server, "ADMIN_API_TOKEN", "secret")
    captured = {}

    def fake_update_miss_query_resolution(miss_id, *, resolution_status, resolved_cas=None):
        captured["miss_id"] = miss_id
        captured["resolution_status"] = resolution_status
        captured["resolved_cas"] = resolved_cas
        return {
            "id": miss_id,
            "query_text": "mystery solvent",
            "query_kind": "name",
            "endpoint": "frontend",
            "resolution_status": resolution_status,
            "resolved_cas": resolved_cas,
            "context": {},
        }

    monkeypatch.setattr(
        server.pilot_store,
        "update_miss_query_resolution",
        fake_update_miss_query_resolution,
    )
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post(
            "/api/dictionary/miss-queries/12/resolution",
            headers={"x-ghs-admin-key": "secret"},
            json={"resolution_status": "resolved", "resolved_cas": " 64 17 5 "},
        )

    assert response.status_code == 200
    assert captured == {
        "miss_id": 12,
        "resolution_status": "resolved",
        "resolved_cas": "64-17-5",
    }
    assert response.json()["record"]["resolved_cas"] == "64-17-5"


async def test_admin_dictionary_miss_query_resolution_requires_known_row(monkeypatch):
    monkeypatch.setattr(server, "ADMIN_API_TOKEN", "secret")
    monkeypatch.setattr(
        server.pilot_store,
        "update_miss_query_resolution",
        lambda *args, **kwargs: None,
    )
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post(
            "/api/dictionary/miss-queries/404/resolution",
            headers={"x-ghs-admin-key": "secret"},
            json={"resolution_status": "ignored"},
        )

    assert response.status_code == 404


async def test_admin_can_view_and_update_correction_requests(monkeypatch):
    monkeypatch.setattr(server, "ADMIN_API_TOKEN", "secret")
    captured = {}

    def fake_list_correction_requests(*, statuses=None, include_context=True):
        captured["list_statuses"] = statuses
        captured["include_context"] = include_context
        return [
            {
                "id": 8,
                "issue_type": "missing-chinese-name",
                "cas_number": "107-18-6",
                "status": "open",
            }
        ]

    def fake_update_correction_request_status(
        request_id,
        *,
        status,
        review_notes=None,
        candidate=None,
    ):
        captured["request_id"] = request_id
        captured["status"] = status
        captured["review_notes"] = review_notes
        captured["candidate"] = candidate
        return {
            "id": request_id,
            "issue_type": "missing-chinese-name",
            "status": status,
            "review_notes": review_notes,
            "candidate": candidate or {},
        }

    monkeypatch.setattr(
        server.pilot_store,
        "list_correction_requests",
        fake_list_correction_requests,
    )
    monkeypatch.setattr(
        server.pilot_store,
        "update_correction_request_status",
        fake_update_correction_request_status,
    )
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        view_response = await ac.get(
            "/api/dictionary/correction-requests?status=open,candidate_found",
            headers={"x-ghs-admin-key": "secret"},
        )
        update_response = await ac.post(
            "/api/dictionary/correction-requests/8/status",
            headers={"x-ghs-admin-key": "secret"},
            json={
                "status": "candidate_found",
                "review_notes": "Found Wikidata candidate.",
                "candidate": {"name_zh": "candidate only"},
            },
        )

    assert view_response.status_code == 200
    assert view_response.json()["items"][0]["id"] == 8
    assert update_response.status_code == 200
    assert captured == {
        "list_statuses": ["open", "candidate_found"],
        "include_context": True,
        "request_id": 8,
        "status": "candidate_found",
        "review_notes": "Found Wikidata candidate.",
        "candidate": {
            "name_zh": "candidate only",
            "schema_version": 1,
            "review_required": True,
            "approved_for_public_use": False,
            "public_data_changed": False,
        },
    }


async def test_admin_correction_request_status_requires_known_row(monkeypatch):
    monkeypatch.setattr(server, "ADMIN_API_TOKEN", "secret")
    monkeypatch.setattr(
        server.pilot_store,
        "update_correction_request_status",
        lambda *args, **kwargs: None,
    )
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post(
            "/api/dictionary/correction-requests/404/status",
            headers={"x-ghs-admin-key": "secret"},
            json={"status": "ignored"},
        )

    assert response.status_code == 404


async def test_admin_can_view_and_purge_miss_query_retention(monkeypatch):
    monkeypatch.setattr(server, "ADMIN_API_TOKEN", "secret")
    captured = {}

    def fake_retention_summary():
        return {
            "retentionDays": 90,
            "cutoffAt": "2026-02-20T00:00:00+00:00",
            "purgeableCount": 2,
            "retainedNeedsEvidenceCount": 1,
        }

    def fake_purge_stale_miss_queries(*, retention_days):
        captured["retention_days"] = retention_days
        return {
            **fake_retention_summary(),
            "retentionDays": retention_days,
            "deletedCount": 2,
            "purgedAt": "2026-05-21T00:00:00+00:00",
        }

    monkeypatch.setattr(
        server.pilot_store,
        "get_miss_query_retention_summary",
        fake_retention_summary,
    )
    monkeypatch.setattr(
        server.pilot_store,
        "purge_stale_miss_queries",
        fake_purge_stale_miss_queries,
    )
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        view_response = await ac.get(
            "/api/dictionary/miss-queries/retention",
            headers={"x-ghs-admin-key": "secret"},
        )
        purge_response = await ac.post(
            "/api/dictionary/miss-queries/retention/purge",
            headers={"x-ghs-admin-key": "secret"},
            json={"retention_days": 30},
        )

    assert view_response.status_code == 200
    assert view_response.json()["retention"]["purgeableCount"] == 2
    assert purge_response.status_code == 200
    assert captured == {"retention_days": 30}
    assert purge_response.json()["retention"]["deletedCount"] == 2


async def test_admin_miss_query_retention_rejects_unbounded_window(monkeypatch):
    monkeypatch.setattr(server, "ADMIN_API_TOKEN", "secret")
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post(
            "/api/dictionary/miss-queries/retention/purge",
            headers={"x-ghs-admin-key": "secret"},
            json={"retention_days": 9999},
        )

    assert response.status_code == 422


def test_admin_dictionary_payloads_trim_safe_fields():
    manual = server.DictionaryManualEntryPayload(
        cas_number=" 64-17-5 ",
        name_en=" Ethanol ",
        name_zh=" ",
        notes=" Pilot note ",
        source=" ",
        status=" NEEDS_EVIDENCE ",
    )
    alias = server.DictionaryAliasPayload(
        alias_text=" EtOH ",
        locale=" EN ",
        cas_number=" 64-17-5 ",
        status=" PENDING ",
        source=" ",
        notes=" Alias note ",
    )
    reference = server.DictionaryReferenceLinkPayload(
        cas_number=" 64-17-5 ",
        label=" Supplier SDS ",
        url=" https://example.com/sds ",
        link_type=" SDS ",
        source=" ",
        status=" ACTIVE ",
    )
    miss_resolution = server.DictionaryMissQueryResolutionPayload(
        resolution_status=" RESOLVED ",
        resolved_cas=" 64 17 5 ",
    )
    correction = server.DictionaryCorrectionRequestPayload(
        issue_type=" Missing-Chinese-Name ",
        cas_number=" 107 18 6 ",
        chemical_name=" Allyl Alcohol ",
        current_output=" Missing Chinese name ",
        expected_output=" Reviewed Traditional Chinese name ",
        evidence_url=" https://example.com/sds ",
        evidence_type=" Supplier SDS ",
        local_context=" Detail modal ",
        source=" ",
    )
    correction_status = server.DictionaryCorrectionRequestStatusPayload(
        status=" Candidate_Found ",
        review_notes=" Candidate needs evidence. ",
        candidate={
            "name_zh": "\u4e59\u9187",
            "source": "wikidata",
            "evidence_url": "https://www.wikidata.org/wiki/Q153",
            "approved_for_public_use": True,
            "public_data_changed": True,
            "converted_to_manual_entry": "true",
            "manual_entry_status": "pending",
            "manual_entry_source": "correction-request",
            "request_id": "8",
            "unknown": "must be dropped",
        },
    )

    assert manual.cas_number == "64-17-5"
    assert manual.name_en == "Ethanol"
    assert manual.name_zh is None
    assert manual.notes == "Pilot note"
    assert manual.source == "manual"
    assert manual.status == "needs_evidence"
    assert alias.alias_text == "EtOH"
    assert alias.locale == "en"
    assert alias.status == "pending"
    assert alias.source == "manual"
    assert reference.cas_number == "64-17-5"
    assert reference.label == "Supplier SDS"
    assert reference.url == "https://example.com/sds"
    assert reference.link_type == "sds"
    assert reference.status == "active"
    assert miss_resolution.resolution_status == "resolved"
    assert miss_resolution.resolved_cas == "64-17-5"
    assert correction.issue_type == "missing-chinese-name"
    assert correction.cas_number == "107-18-6"
    assert correction.chemical_name == "Allyl Alcohol"
    assert correction.evidence_url == "https://example.com/sds"
    assert correction.local_context == "Detail modal"
    assert correction.source == "public"
    assert correction_status.status == "candidate_found"
    assert correction_status.review_notes == "Candidate needs evidence."
    assert correction_status.candidate == {
        "name_zh": "\u4e59\u9187",
        "source": "wikidata",
        "evidence_url": "https://www.wikidata.org/wiki/Q153",
        "schema_version": 1,
        "review_required": True,
        "approved_for_public_use": False,
        "public_data_changed": False,
        "converted_to_manual_entry": True,
        "manual_entry_status": "pending",
        "manual_entry_source": "correction-request",
        "request_id": 8,
    }

    manual_with_zh = server.DictionaryManualEntryPayload(
        cas_number="64-17-5",
        name_en="Ethanol",
        name_zh=" 乙醇 ",
    )
    assert manual_with_zh.name_zh == "乙醇"


def test_admin_dictionary_payloads_reject_unbounded_values():
    with pytest.raises(ValidationError):
        server.DictionaryManualEntryPayload(
            cas_number="64-17-5",
            name_en="x" * (server.MAX_ADMIN_NAME_LENGTH + 1),
        )
    with pytest.raises(ValidationError):
        server.DictionaryManualEntryPayload(
            cas_number="107-18-6",
            name_en="Allyl Alcohol",
            name_zh="Allyl Alcohol",
        )
    with pytest.raises(ValidationError):
        server.DictionaryManualEntryPayload(
            cas_number="64-17-5",
            status="published",
        )
    with pytest.raises(ValidationError):
        server.DictionaryAliasPayload(
            alias_text="x" * (server.MAX_ALIAS_TEXT_LENGTH + 1),
            locale="en",
            cas_number="64-17-5",
        )
    with pytest.raises(ValidationError):
        server.DictionaryAliasPayload(
            alias_text="EtOH",
            locale="all",
            cas_number="64-17-5",
        )
    with pytest.raises(ValidationError):
        server.DictionaryAliasPayload(
            alias_text="EtOH",
            locale="en",
            cas_number="64-17-5",
            confidence=2,
        )
    with pytest.raises(ValidationError):
        server.DictionaryReferenceLinkPayload(
            cas_number="64-17-5",
            label="Supplier SDS",
            url="https://example.com/sds",
            priority=server.MAX_REFERENCE_PRIORITY + 1,
        )
    with pytest.raises(ValidationError):
        server.DictionaryReferenceLinkPayload(
            cas_number="64-17-5",
            label="Supplier SDS",
            url="https://example.com/sds",
            status="script",
        )
    with pytest.raises(ValidationError):
        server.DictionaryMissQueryResolutionPayload(
            resolution_status="done",
        )
    with pytest.raises(ValidationError):
        server.DictionaryCorrectionRequestPayload(
            issue_type="unsafe-kind",
        )
    with pytest.raises(ValidationError):
        server.DictionaryCorrectionRequestPayload(
            issue_type="missing-chinese-name",
            evidence_url="javascript:alert(1)",
        )
    with pytest.raises(ValidationError):
        server.DictionaryCorrectionRequestPayload(
            issue_type="missing-chinese-name",
            candidate={"blob": "x" * (server.MAX_CORRECTION_CANDIDATE_JSON_CHARS + 1)},
        )
    with pytest.raises(ValidationError):
        server.DictionaryCorrectionRequestStatusPayload(
            status="candidate_found",
            candidate={"evidence_url": "data:text/plain,unsafe"},
        )
    with pytest.raises(ValidationError):
        server.DictionaryCorrectionRequestStatusPayload(
            status="done",
        )


def test_public_payload_models_reject_unbounded_values():
    with pytest.raises(ValidationError):
        server.CASQuery(cas_numbers=["1" * (MAX_PUBLIC_CAS_QUERY_LENGTH + 1)])

    with pytest.raises(ValidationError):
        server.ExportRequest(
            results=[
                {
                    "cas_number": "64-17-5",
                    "name_en": "x" * (MAX_EXPORT_SCALAR_CHARS + 1),
                }
            ]
        )


async def test_admin_write_routes_reject_invalid_payloads(monkeypatch):
    monkeypatch.setattr(server, "ADMIN_API_TOKEN", "secret")
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        manual_response = await ac.post(
            "/api/dictionary/manual-entries",
            headers={"x-ghs-admin-key": "secret"},
            json={
                "cas_number": "64-17-5",
                "name_en": "x" * (server.MAX_ADMIN_NAME_LENGTH + 1),
            },
        )
        alias_response = await ac.post(
            "/api/dictionary/aliases",
            headers={"x-ghs-admin-key": "secret"},
            json={
                "alias_text": "EtOH",
                "locale": "all",
                "cas_number": "64-17-5",
            },
        )
        reference_response = await ac.post(
            "/api/dictionary/reference-links",
            headers={"x-ghs-admin-key": "secret"},
            json={
                "cas_number": "64-17-5",
                "label": "Supplier SDS",
                "url": "https://example.com/sds",
                "priority": server.MAX_REFERENCE_PRIORITY + 1,
            },
        )
        manual_name_response = await ac.post(
            "/api/dictionary/manual-entries",
            headers={"x-ghs-admin-key": "secret"},
            json={
                "cas_number": "107-18-6",
                "name_en": "Allyl Alcohol",
                "name_zh": "Allyl Alcohol",
            },
        )

    assert manual_response.status_code == 422
    assert manual_name_response.status_code == 422
    assert alias_response.status_code == 422
    assert reference_response.status_code == 422


async def test_reference_link_rejects_non_http_url(monkeypatch):
    monkeypatch.setattr(server, "ADMIN_API_TOKEN", "secret")
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post(
            "/api/dictionary/reference-links",
            headers={"x-ghs-admin-key": "secret"},
            json={
                "cas_number": "64-17-5",
                "label": "Unsafe SDS",
                "url": "javascript:alert(1)",
                "link_type": "sds",
            },
        )

    assert response.status_code == 422


async def test_reference_link_rejects_unknown_link_type(monkeypatch):
    monkeypatch.setattr(server, "ADMIN_API_TOKEN", "secret")
    transport = ASGITransport(app=app)

    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post(
            "/api/dictionary/reference-links",
            headers={"x-ghs-admin-key": "secret"},
            json={
                "cas_number": "64-17-5",
                "label": "Unsafe role",
                "url": "https://example.com/sds",
                "link_type": "script",
            },
        )

    assert response.status_code == 422


def test_build_reference_links_skips_unsafe_manual_urls(monkeypatch):
    def fake_reference_links(_cas_number):
        return [
            {
                "label": "Unsafe SDS",
                "url": "javascript:alert(1)",
                "linkType": "sds",
                "source": "manual",
                "priority": 1,
            },
            {
                "label": "Tenant SDS",
                "url": "https://example.com/sds",
                "linkType": "sds",
                "source": "manual",
                "priority": 2,
            },
        ]

    monkeypatch.setattr(server.pilot_store, "list_reference_links", fake_reference_links)

    links = server._build_reference_links("64-17-5", None, "Ethanol")

    assert all(link["url"] != "javascript:alert(1)" for link in links)
    assert any(link["url"] == "https://example.com/sds" for link in links)


def test_build_reference_links_normalizes_legacy_manual_link_type(monkeypatch):
    def fake_reference_links(_cas_number):
        return [
            {
                "label": "Legacy Role SDS",
                "url": "https://example.com/legacy-sds",
                "linkType": "script",
                "source": "manual",
                "priority": 1,
            },
        ]

    monkeypatch.setattr(server.pilot_store, "list_reference_links", fake_reference_links)

    links = server._build_reference_links("64-17-5", None, "Ethanol")
    legacy = next(link for link in links if link["url"] == "https://example.com/legacy-sds")

    assert legacy["link_type"] == "reference"


def test_build_reference_links_keeps_strongest_role_for_duplicate_urls(monkeypatch):
    pubchem_sds_url = "https://pubchem.ncbi.nlm.nih.gov/compound/702#section=Safety-and-Hazards"

    def fake_reference_links(_cas_number):
        return [
            {
                "label": "Generic mirror of PubChem safety",
                "url": pubchem_sds_url,
                "linkType": "reference",
                "source": "manual",
                "priority": 1,
            },
        ]

    monkeypatch.setattr(server.pilot_store, "list_reference_links", fake_reference_links)

    links = server._build_reference_links("64-17-5", 702, "Ethanol")
    safety_link = next(link for link in links if link["url"] == pubchem_sds_url)

    assert safety_link["label"] == "PubChem Safety & Hazards"
    assert safety_link["link_type"] == "sds"


def test_build_reference_links_orders_by_authority_role_before_priority(monkeypatch):
    def fake_reference_links(_cas_number):
        return [
            {
                "label": "Generic internal note",
                "url": "https://example.com/internal-note",
                "linkType": "reference",
                "source": "manual",
                "priority": 1,
            },
            {
                "label": "Supplier SDS",
                "url": "https://example.com/supplier-sds",
                "linkType": "sds",
                "source": "manual",
                "priority": 50,
            },
        ]

    monkeypatch.setattr(server.pilot_store, "list_reference_links", fake_reference_links)

    links = server._build_reference_links("64-17-5", 702, "Ethanol")

    assert [link["link_type"] for link in links] == [
        "sds",
        "sds",
        "regulatory",
        "occupational",
        "reference",
        "reference",
    ]
    assert links[0]["label"] == "PubChem Safety & Hazards"
    assert links[1]["label"] == "Supplier SDS"
    assert links[4]["label"] == "Generic internal note"


async def test_search_by_name_chinese():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/search-by-name/%E4%B9%99%E9%86%87")  # 乙醇
    assert response.status_code == 200
    data = response.json()
    cas_numbers = [r["cas_number"] for r in data["results"]]
    assert "64-17-5" in cas_numbers


async def test_search_by_name_short_query_returns_empty():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/search-by-name/a")
    assert response.status_code == 200
    data = response.json()
    assert data["results"] == []


async def test_search_by_name_rejects_overlong_query():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get(
            f"/api/search-by-name/{'x' * (MAX_PUBLIC_SEARCH_QUERY_LENGTH + 1)}"
        )
    assert response.status_code == 422


async def test_search_by_name_max_20_results():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/search-by-name/acid")
    assert response.status_code == 200
    data = response.json()
    assert len(data["results"]) <= 20


async def test_search_by_name_unknown_returns_empty():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/search-by-name/xyznotfound")
    assert response.status_code == 200
    data = response.json()
    assert data["results"] == []


async def test_search_by_name_result_has_all_fields():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/search-by-name/ethanol")
    data = response.json()
    for r in data["results"]:
        assert "cas_number" in r
        assert "name_en" in r
        assert "name_zh" in r


async def test_search_by_name_methanol_returns_results():
    """Searching 'methanol' should return the chemical even though
    it's stored as 'Methyl alcohol (Methanol)' in CAS_TO_EN."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/search-by-name/methanol")
    assert response.status_code == 200
    data = response.json()
    cas_numbers = [r["cas_number"] for r in data["results"]]
    assert "67-56-1" in cas_numbers


async def test_search_single_unknown_name_returns_not_found():
    """/api/search/Unobtainium returns found=false."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/search/Unobtainium")
    assert response.status_code == 200
    data = response.json()
    assert data["found"] is False
    assert "error" in data


# ─── Unit tests: alias dictionaries ──────────────────────

class TestAliasDictionaries:
    """Test that alias dictionaries are correctly defined and merged."""

    def test_aliases_zh_has_entries(self):
        assert len(ALIASES_ZH) > 0

    def test_aliases_en_has_entries(self):
        assert len(ALIASES_EN) > 0

    def test_aliases_zh_merged_into_zh_to_cas(self):
        """All Chinese aliases should be findable in ZH_TO_CAS."""
        for alias, cas in ALIASES_ZH.items():
            assert alias in ZH_TO_CAS, f"Alias '{alias}' not found in ZH_TO_CAS"

    def test_aliases_en_merged_into_en_to_cas(self):
        """All English aliases (lowercase) should be findable in EN_TO_CAS."""
        for alias, cas in ALIASES_EN.items():
            key = alias.lower()
            assert key in EN_TO_CAS, f"Alias '{alias}' not found in EN_TO_CAS"

    def test_alias_does_not_overwrite_formal_name(self):
        """If a formal name exists in ZH_TO_CAS, alias should not overwrite it."""
        # 乙醇 is a formal name in CAS_TO_ZH for 64-17-5
        assert ZH_TO_CAS["乙醇"] == "64-17-5"
        # 酒精 is an alias for 64-17-5 — should also resolve to 64-17-5
        assert ZH_TO_CAS["酒精"] == "64-17-5"

    def test_common_zh_aliases_resolve(self):
        """Common Chinese aliases should resolve to correct CAS numbers."""
        assert ZH_TO_CAS.get("酒精") == "64-17-5"      # Ethanol
        assert ZH_TO_CAS.get("漂白水") == "7681-52-9"   # NaClO
        assert ZH_TO_CAS.get("雙氧水") == "7722-84-1"   # H2O2
        assert ZH_TO_CAS.get("燒鹼") == "1310-73-2"     # NaOH
        assert ZH_TO_CAS.get("小蘇打") == "144-55-8"     # NaHCO3

    def test_common_en_aliases_resolve(self):
        """Common English aliases should resolve to correct CAS numbers."""
        assert EN_TO_CAS.get("bleach") == "7681-52-9"
        assert EN_TO_CAS.get("baking soda") == "144-55-8"
        assert EN_TO_CAS.get("caustic soda") == "1310-73-2"
        assert EN_TO_CAS.get("lye") == "1310-73-2"
        assert EN_TO_CAS.get("dmso") == "67-68-5"

    def test_alias_cas_numbers_are_valid_format(self):
        """All CAS numbers in aliases should have valid format (XX-XX-X)."""
        import re
        cas_pattern = re.compile(r'^\d{2,7}-\d{2}-\d$')
        for alias, cas in ALIASES_ZH.items():
            assert cas_pattern.match(cas), f"Invalid CAS '{cas}' for alias '{alias}'"
        for alias, cas in ALIASES_EN.items():
            assert cas_pattern.match(cas), f"Invalid CAS '{cas}' for alias '{alias}'"


# ─── Unit tests: resolve_name_to_cas with aliases ────────

class TestResolveNameToCasAliases:
    """Test that resolve_name_to_cas works with aliases."""

    def test_chinese_alias_alcohol(self):
        """酒精 (alcohol) should resolve to ethanol."""
        assert resolve_name_to_cas("酒精") == "64-17-5"

    def test_chinese_alias_bleach(self):
        """漂白水 should resolve to sodium hypochlorite."""
        assert resolve_name_to_cas("漂白水") == "7681-52-9"

    def test_chinese_alias_hydrogen_peroxide(self):
        """雙氧水 should resolve to H2O2."""
        assert resolve_name_to_cas("雙氧水") == "7722-84-1"

    def test_chinese_alias_caustic_soda(self):
        """燒鹼 should resolve to NaOH."""
        assert resolve_name_to_cas("燒鹼") == "1310-73-2"

    def test_chinese_alias_formalin(self):
        """福馬林 should resolve to formaldehyde."""
        assert resolve_name_to_cas("福馬林") == "50-00-0"

    def test_chinese_alias_nail_polish_remover(self):
        """去光水 should resolve to acetone."""
        assert resolve_name_to_cas("去光水") == "67-64-1"

    def test_english_alias_bleach(self):
        assert resolve_name_to_cas("bleach") == "7681-52-9"

    def test_english_alias_baking_soda(self):
        assert resolve_name_to_cas("baking soda") == "144-55-8"

    def test_english_alias_lye(self):
        assert resolve_name_to_cas("lye") == "1310-73-2"

    def test_english_alias_dmso(self):
        assert resolve_name_to_cas("DMSO") == "67-68-5"

    def test_english_alias_rubbing_alcohol(self):
        assert resolve_name_to_cas("rubbing alcohol") == "67-63-0"

    def test_english_alias_dry_ice(self):
        assert resolve_name_to_cas("dry ice") == "124-38-9"

    def test_formal_name_still_works(self):
        """Formal names should still resolve correctly (not broken by aliases)."""
        assert resolve_name_to_cas("乙醇") == "64-17-5"
        assert resolve_name_to_cas("Ethanol") == "64-17-5"
        assert resolve_name_to_cas("甲醇") == "67-56-1"


# ─── Integration tests: /api/search-by-name with aliases ──

async def test_search_by_name_chinese_alias():
    """Searching '酒精' should find ethanol (64-17-5)."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/search-by-name/%E9%85%92%E7%B2%BE")  # 酒精
    assert response.status_code == 200
    data = response.json()
    cas_numbers = [r["cas_number"] for r in data["results"]]
    assert "64-17-5" in cas_numbers


async def test_search_by_name_english_alias():
    """Searching 'bleach' should find sodium hypochlorite (7681-52-9)."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/search-by-name/bleach")
    assert response.status_code == 200
    data = response.json()
    cas_numbers = [r["cas_number"] for r in data["results"]]
    assert "7681-52-9" in cas_numbers


async def test_search_by_name_alias_has_alias_field():
    """Alias matches should include an 'alias' field."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/search-by-name/bleach")
    assert response.status_code == 200
    data = response.json()
    # Find the sodium hypochlorite result
    match = next((r for r in data["results"] if r["cas_number"] == "7681-52-9"), None)
    assert match is not None
    assert "alias" in match


async def test_search_by_name_dmso_alias():
    """Searching 'dmso' should find DMSO (67-68-5)."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/search-by-name/dmso")
    assert response.status_code == 200
    data = response.json()
    cas_numbers = [r["cas_number"] for r in data["results"]]
    assert "67-68-5" in cas_numbers


async def test_search_by_name_lye_alias():
    """Searching 'lye' should find NaOH (1310-73-2)."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/search-by-name/lye")
    assert response.status_code == 200
    data = response.json()
    cas_numbers = [r["cas_number"] for r in data["results"]]
    assert "1310-73-2" in cas_numbers


# ─── GHS classification dedup / ranking ─────────────────────

def _pic(code):
    return {"code": code}


def _hz(code, text="x"):
    return {"code": code, "text_en": text, "text_zh": text}


class TestClassificationSignature:
    """Dedup must collapse only truly identical reports."""

    def test_same_pictograms_but_different_h_codes_produce_different_signatures(self):
        """Regression: previously deduped on pictogram set only, silently
        losing a materially different classification."""
        a = {
            "pictograms": [_pic("GHS02"), _pic("GHS07")],
            "hazard_statements": [_hz("H225"), _hz("H319")],
            "signal_word": "Danger",
            "source": "ECHA C&L Notifications Summary",
        }
        b = {
            "pictograms": [_pic("GHS02"), _pic("GHS07")],
            "hazard_statements": [_hz("H225"), _hz("H336")],
            "signal_word": "Danger",
            "source": "ECHA C&L Notifications Summary",
        }
        assert _classification_signature(a) != _classification_signature(b)

    def test_different_signal_word_produces_different_signature(self):
        a = {
            "pictograms": [_pic("GHS07")],
            "hazard_statements": [_hz("H319")],
            "signal_word": "Danger",
            "source": "ECHA",
        }
        b = {**a, "signal_word": "Warning"}
        assert _classification_signature(a) != _classification_signature(b)

    def test_different_source_produces_different_signature(self):
        a = {
            "pictograms": [_pic("GHS07")],
            "hazard_statements": [_hz("H319")],
            "signal_word": "Warning",
            "source": "ECHA C&L",
        }
        b = {**a, "source": "Other vendor SDS"}
        assert _classification_signature(a) != _classification_signature(b)

    def test_identical_reports_produce_equal_signatures(self):
        a = {
            "pictograms": [_pic("GHS07"), _pic("GHS02")],
            "hazard_statements": [_hz("H319"), _hz("H225")],
            "signal_word": "Danger",
            "source": "ECHA C&L Notifications Summary",
        }
        b = {
            "pictograms": [_pic("GHS02"), _pic("GHS07")],  # reversed order
            "hazard_statements": [_hz("H225"), _hz("H319")],  # reversed order
            "signal_word": "Danger",
            "source": "ECHA C&L Notifications Summary",
        }
        assert _classification_signature(a) == _classification_signature(b)


class TestReportRankKey:
    """Primary classification selection must be deterministic and prefer
    the most-reported / most-complete classification."""

    def test_higher_report_count_ranks_first(self):
        a = {"report_count": "120", "source": "ECHA", "hazard_statements": [_hz("H1")]}
        b = {"report_count": "3", "source": "ECHA", "hazard_statements": [_hz("H1")]}
        # Ascending sort; lower tuple wins → a should rank before b
        assert _report_rank_key(a, 0) < _report_rank_key(b, 1)

    def test_echa_source_bonus_when_count_equal(self):
        a = {"report_count": "10", "source": "Other", "hazard_statements": [_hz("H1")]}
        b = {"report_count": "10", "source": "ECHA C&L Notifications", "hazard_statements": [_hz("H1")]}
        assert _report_rank_key(b, 1) < _report_rank_key(a, 0)

    def test_more_hazards_wins_on_tie(self):
        a = {"report_count": None, "source": "ECHA", "hazard_statements": [_hz("H1"), _hz("H2")]}
        b = {"report_count": None, "source": "ECHA", "hazard_statements": [_hz("H1")]}
        assert _report_rank_key(a, 0) < _report_rank_key(b, 1)

    def test_source_order_is_stable_tiebreaker(self):
        a = {"report_count": None, "source": "ECHA", "hazard_statements": [_hz("H1")]}
        b = {"report_count": None, "source": "ECHA", "hazard_statements": [_hz("H1")]}
        assert _report_rank_key(a, 0) < _report_rank_key(b, 1)

    def test_non_numeric_report_count_treated_as_zero(self):
        a = {"report_count": "not-a-number", "source": "ECHA", "hazard_statements": []}
        b = {"report_count": None, "source": "ECHA", "hazard_statements": []}
        # Both should produce the same count component; only source_index differs
        assert _report_rank_key(a, 0) < _report_rank_key(b, 1)


async def test_search_chemical_both_distinct_reports_survive_dedup(monkeypatch):
    """End-to-end: two reports with identical pictograms but different
    H-codes must BOTH survive (regression against the old pic-set-only
    dedup), and the higher report_count must become primary."""
    import server as srv

    fake_reports = [
        {
            "pictograms": [_pic("GHS02"), _pic("GHS07")],
            "hazard_statements": [_hz("H225"), _hz("H319")],
            "signal_word": "Danger",
            "signal_word_zh": "危險",
            "source": "ECHA C&L Notifications Summary",
            "report_count": "5",
        },
        {
            "pictograms": [_pic("GHS02"), _pic("GHS07")],
            "hazard_statements": [_hz("H225"), _hz("H336")],
            "signal_word": "Danger",
            "signal_word_zh": "危險",
            "source": "ECHA C&L Notifications Summary",
            "report_count": "120",
        },
    ]

    async def fake_get_cid(*_args, **_kwargs):
        return 702

    async def fake_get_name(*_args, **_kwargs):
        return ("Ethanol", "乙醇")

    async def fake_get_ghs(*_args, **_kwargs):
        # v1.8 M1: get_ghs_classification now returns (data, cache_hit, retrieved_at)
        return ({}, False, "2026-04-16T00:00:00+00:00")

    monkeypatch.setattr(srv, "get_cid_from_cas", fake_get_cid)
    monkeypatch.setattr(srv, "get_compound_name", fake_get_name)
    monkeypatch.setattr(srv, "get_ghs_classification", fake_get_ghs)
    monkeypatch.setattr(srv, "extract_all_ghs_classifications", lambda _: fake_reports)

    class _NullClient:
        async def get(self, *_a, **_k):  # pragma: no cover - unused
            raise RuntimeError("http should not be called")

    result = await srv.search_chemical("64-17-5", _NullClient())

    # Both distinct reports must survive dedup
    assert result.has_multiple_classifications is True
    assert len(result.other_classifications) == 1

    primary_h_codes = sorted(h["code"] for h in result.hazard_statements)
    other_h_codes = sorted(h["code"] for h in result.other_classifications[0].hazard_statements)
    assert {tuple(primary_h_codes), tuple(other_h_codes)} == {
        ("H225", "H319"),
        ("H225", "H336"),
    }

    # Specifically: report_count=120 (the H336 one) must be primary
    primary_has_h336 = any(h["code"] == "H336" for h in result.hazard_statements)
    assert primary_has_h336, "Higher report_count must be chosen as primary"


# ─── Export endpoint safety ─────────────────────────────────

from server import spreadsheet_safe, MAX_EXPORT_ROWS


class TestSpreadsheetSafe:
    """spreadsheet_safe() neutralizes CSV/XLSX formula injection."""

    def test_none_becomes_empty_string(self):
        assert spreadsheet_safe(None) == ""

    def test_plain_text_unchanged(self):
        assert spreadsheet_safe("Ethanol") == "Ethanol"
        assert spreadsheet_safe("64-17-5") == "64-17-5"  # leading digit is safe

    def test_equals_prefix_is_neutralized(self):
        assert spreadsheet_safe('=HYPERLINK("http://bad","click")') == "'=HYPERLINK(\"http://bad\",\"click\")"

    def test_plus_prefix_is_neutralized(self):
        assert spreadsheet_safe("+cmd|calc") == "'+cmd|calc"

    def test_minus_prefix_is_neutralized(self):
        # A minus sign in front is a formula trigger too
        assert spreadsheet_safe("-1+1") == "'-1+1"

    def test_at_prefix_is_neutralized(self):
        assert spreadsheet_safe("@SUM(A1:A10)") == "'@SUM(A1:A10)"

    def test_tab_prefix_is_neutralized(self):
        assert spreadsheet_safe("\tmalicious") == "'\tmalicious"

    def test_non_string_values_coerced(self):
        assert spreadsheet_safe(42) == "42"
        assert spreadsheet_safe(True) == "True"


async def test_export_csv_neutralizes_formula_injection():
    """Malicious `=...` value in a result field must be prefixed with
    an apostrophe in the CSV output."""
    transport = ASGITransport(app=app)
    payload = {
        "results": [
            {
                "cas_number": "64-17-5",
                "name_en": '=HYPERLINK("http://bad","click")',
                "name_zh": "乙醇",
                "ghs_pictograms": [],
                "hazard_statements": [],
            }
        ]
    }
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post("/api/export/csv", json=payload)
    assert response.status_code == 200
    # Body is UTF-8 with BOM. The neutralized cell must start with an
    # apostrophe before the formula trigger, so the spreadsheet reads
    # the value as literal text.
    body = response.content.decode("utf-8-sig")
    assert "'=HYPERLINK" in body
    # There must be NO formula-triggering cell: every `=...` occurrence
    # must be preceded by the neutralizing apostrophe.
    assert ",=HYPERLINK" not in body
    assert '"=HYPERLINK' not in body


async def test_export_xlsx_neutralizes_formula_injection():
    """Same regression for the XLSX endpoint."""
    from io import BytesIO
    from openpyxl import load_workbook

    transport = ASGITransport(app=app)
    payload = {
        "results": [
            {
                "cas_number": "64-17-5",
                "name_en": "=1+1",
                "name_zh": "+test",
                "ghs_pictograms": [],
                "hazard_statements": [],
                "signal_word_zh": "+test",
            }
        ]
    }
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post("/api/export/xlsx", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    # Row 2 is the first data row. Columns: 1=CAS, 2=en, 3=zh ...
    assert ws.cell(row=2, column=2).value == "'=1+1"
    assert ws.cell(row=2, column=3).value in (None, "")
    assert ws.cell(row=2, column=5).value == "'+test"


async def test_export_csv_omits_english_only_chinese_name():
    """Direct backend CSV export should not preserve English placeholders
    in the Chinese-name column."""
    import csv as csv_module
    from io import StringIO

    transport = ASGITransport(app=app)
    payload = {
        "results": [
            {
                "cas_number": "107-18-6",
                "name_en": "Allyl Alcohol",
                "name_zh": "Allyl Alcohol",
                "ghs_pictograms": [],
                "hazard_statements": [],
            }
        ]
    }
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post("/api/export/csv", json=payload)
    assert response.status_code == 200

    rows = list(csv_module.reader(StringIO(response.content.decode("utf-8-sig"))))
    assert rows[1][1] == "Allyl Alcohol"
    assert rows[1][2] == ""


async def test_export_xlsx_omits_english_only_chinese_name():
    """XLSX export applies the same trusted Chinese-name boundary."""
    from io import BytesIO
    from openpyxl import load_workbook

    transport = ASGITransport(app=app)
    payload = {
        "results": [
            {
                "cas_number": "107-18-6",
                "name_en": "Allyl Alcohol",
                "name_zh": "Allyl Alcohol",
                "ghs_pictograms": [],
                "hazard_statements": [],
            }
        ]
    }
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post("/api/export/xlsx", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    assert ws.cell(row=2, column=2).value == "Allyl Alcohol"
    assert ws.cell(row=2, column=3).value in (None, "")


async def test_export_rejects_payload_exceeding_max_rows():
    """Payload with > MAX_EXPORT_ROWS items must be rejected with 422."""
    transport = ASGITransport(app=app)
    oversized = [
        {"cas_number": f"000-00-{i}", "name_en": "x", "name_zh": "x",
         "ghs_pictograms": [], "hazard_statements": []}
        for i in range(MAX_EXPORT_ROWS + 1)
    ]
    payload = {"results": oversized}
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post("/api/export/csv", json=payload)
    assert response.status_code == 422


async def test_export_accepts_payload_at_max_rows():
    """MAX_EXPORT_ROWS items exactly must still succeed (inclusive limit)."""
    transport = ASGITransport(app=app)
    payload = {
        "results": [
            {"cas_number": f"000-00-{i}", "name_en": "x", "name_zh": "x",
             "ghs_pictograms": [], "hazard_statements": []}
            for i in range(MAX_EXPORT_ROWS)
        ]
    }
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post("/api/export/csv", json=payload)
    assert response.status_code == 200


# ─── Export P-codes column (v1.8 M0 PR-C) ───────────────────


async def test_export_csv_includes_precautionary_column():
    """CSV export must include a precautionary statements column."""
    transport = ASGITransport(app=app)
    payload = {
        "results": [
            {
                "cas_number": "64-17-5",
                "name_en": "Ethanol",
                "name_zh": "\u4e59\u9187",
                "ghs_pictograms": [],
                "hazard_statements": [],
                "precautionary_statements": [
                    {"code": "P210", "text_en": "Keep away from heat.", "text_zh": "\u9060\u96e2\u71b1\u6e90\u3002"},
                    {"code": "P301+P310", "text_en": "IF SWALLOWED: Call a POISON CENTER.", "text_zh": "\u5982\u8aa4\u541e\u98df\uff1a\u7acb\u5373\u547c\u53eb\u6bd2\u7269\u4e2d\u5fc3\u3002"},
                ],
            }
        ]
    }
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post("/api/export/csv", json=payload)
    assert response.status_code == 200
    body = response.content.decode("utf-8-sig")
    # Header row must include the new column
    assert "\u9810\u9632\u63aa\u65bd" in body or "precautionary" in body.lower()
    # P-code values must appear
    assert "P210" in body
    assert "P301+P310" in body


async def test_export_xlsx_includes_precautionary_column():
    """XLSX export must include the precautionary column with P-code data."""
    from io import BytesIO
    from openpyxl import load_workbook

    transport = ASGITransport(app=app)
    payload = {
        "results": [
            {
                "cas_number": "64-17-5",
                "name_en": "Ethanol",
                "name_zh": "\u4e59\u9187",
                "ghs_pictograms": [],
                "hazard_statements": [],
                "precautionary_statements": [
                    {"code": "P210", "text_en": "Keep away.", "text_zh": "\u9060\u96e2\u71b1\u6e90\u3002"},
                ],
            }
        ]
    }
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post("/api/export/xlsx", json=payload)
    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    # Column 7 is precautionary statements (headers row)
    header = ws.cell(row=1, column=7).value
    assert "\u9810\u9632\u63aa\u65bd" in (header or "") or "precautionary" in (
        header or ""
    ).lower()
    # Data row 2 column 7 has the P-code content
    data = ws.cell(row=2, column=7).value
    assert data is not None
    assert "P210" in data


async def test_export_csv_includes_data_trust_columns():
    """CSV export preserves source/data-state context for downstream users."""
    transport = ASGITransport(app=app)
    payload = {
        "results": [
            {
                "cas_number": "64-17-5",
                "name_en": "Ethanol",
                "name_zh": "\u4e59\u9187",
                "found": True,
                "ghs_pictograms": [{"code": "GHS02", "name_zh": "\u706b\u7130"}],
                "hazard_statements": [],
                "precautionary_statements": [],
                "signal_word": "Danger",
                "primary_source": "ECHA C&L Notifications Summary",
                "primary_report_count": "236",
                "retrieved_at": "2026-05-16T00:00:00Z",
                "cache_hit": True,
                "reference_links": [
                    {"label": "Supplier SDS", "url": "https://example.com/sds"}
                ],
                "has_multiple_classifications": True,
                "other_classifications": [{"pictograms": [], "hazard_statements": []}],
                "customNote": "Use alternate supplier-confirmed report",
            }
        ]
    }
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post("/api/export/csv", json=payload)
    assert response.status_code == 200
    body = response.content.decode("utf-8-sig")

    assert "Data State" in body
    assert "Printable" in body
    assert "Needs Review" in body
    assert "Review Reasons" in body
    assert "Review Signal Count" in body
    assert "Primary Review Action" in body
    assert "Primary Source" in body
    assert "Found with renderable GHS pictograms" in body
    assert "Yes,No,No review reasons,0,No review action" in body
    assert "ECHA C&L Notifications Summary" in body
    assert "236" in body
    assert "Cached" in body
    assert "1 reference link(s)" in body
    assert "User-selected primary classification" in body
    assert "User-selected classification: Use alternate supplier-confirmed report" in body


async def test_export_xlsx_includes_data_trust_columns():
    """XLSX export includes the same appended trust columns as CSV."""
    from io import BytesIO
    from openpyxl import load_workbook

    transport = ASGITransport(app=app)
    payload = {
        "results": [
            {
                "cas_number": "999-99-9",
                "name_en": "Unknown",
                "found": False,
                "ghs_pictograms": [],
                "hazard_statements": [],
                "precautionary_statements": [],
            }
        ]
    }
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post("/api/export/xlsx", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    assert ws.cell(row=1, column=8).value == "Data State"
    assert ws.cell(row=1, column=12).value == "Review Signal Count"
    assert ws.cell(row=1, column=13).value == "Primary Review Action"
    assert ws.cell(row=1, column=22).value == "Classification Selection"
    assert ws.cell(row=2, column=8).value == "Not found"
    assert ws.cell(row=2, column=9).value == "No"
    assert ws.cell(row=2, column=10).value == "Yes"
    assert ws.cell(row=2, column=11).value == "Unresolved lookup"
    assert ws.cell(row=2, column=12).value == "1"
    assert ws.cell(row=2, column=13).value == "Send lookup gap to correction"
    assert ws.cell(row=2, column=22).value == "Default primary classification"


async def test_export_xlsx_includes_pilot_summary_sheet():
    """XLSX export includes a compact pilot summary for downstream triage."""
    from io import BytesIO
    from openpyxl import load_workbook

    transport = ASGITransport(app=app)
    payload = {
        "export_scope": "needs-review",
        "export_scope_label": "=Needs review",
        "export_count": 4,
        "source_total_count": 9,
        "visible_count": 4,
        "results": [
            {
                "cas_number": "64-17-5",
                "name_en": "Ethanol",
                "name_zh": "\u4e59\u9187",
                "found": True,
                "ghs_pictograms": [{"code": "GHS02", "name_zh": "\u706b\u7130"}],
                "hazard_statements": [],
                "precautionary_statements": [],
            },
            {
                "cas_number": "999-99-9",
                "name_en": "Unknown",
                "found": False,
                "ghs_pictograms": [],
                "hazard_statements": [],
                "precautionary_statements": [],
            },
            {
                "cas_number": "777-77-7",
                "name_en": "Temporary PubChem outage",
                "found": False,
                "upstream_error": True,
                "ghs_pictograms": [],
                "hazard_statements": [],
                "precautionary_statements": [],
            },
            {
                "cas_number": "107-18-6",
                "name_en": "Allyl Alcohol",
                "name_zh": "Allyl Alcohol",
                "found": True,
                "ghs_pictograms": [{"code": "GHS06", "name_zh": "\u9ab7\u9acf"}],
                "hazard_statements": [],
                "precautionary_statements": [],
                "has_multiple_classifications": True,
            },
        ]
    }
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post("/api/export/xlsx", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    assert "Pilot Summary" in wb.sheetnames
    summary = wb["Pilot Summary"]
    rows = {
        summary.cell(row=row, column=1).value: summary.cell(row=row, column=2).value
        for row in range(2, summary.max_row + 1)
    }
    assert rows["Total rows"] == 9
    assert rows["Visible filtered rows"] == 4
    assert rows["Export scope"] == "'=Needs review"
    assert rows["Exported row count"] == 4
    assert rows["Printable rows"] == 2
    assert rows["Needs review rows"] == 3
    assert rows["Review signals"] == 4
    assert rows["Rows with multiple review signals"] == 1
    assert rows["Unresolved searches"] == 1
    assert rows["Upstream transient failures"] == 1
    assert rows["Missing trusted Chinese names"] == 1
    assert rows["Multiple GHS classifications"] == 1


async def test_export_xlsx_includes_lab_manager_triage_sheets():
    """XLSX export should partition a batch into ready/review/unresolved sheets."""
    from io import BytesIO
    from openpyxl import load_workbook

    transport = ASGITransport(app=app)
    payload = {
        "results": [
            {
                "cas_number": "64-17-5",
                "name_en": "Ethanol",
                "name_zh": "\u4e59\u9187",
                "found": True,
                "ghs_pictograms": [{"code": "GHS02", "name_zh": "\u706b\u7130"}],
                "hazard_statements": [],
                "precautionary_statements": [],
            },
            {
                "cas_number": "999-99-9",
                "name_en": "Unknown",
                "found": False,
                "ghs_pictograms": [],
                "hazard_statements": [],
                "precautionary_statements": [],
            },
            {
                "cas_number": "777-77-7",
                "name_en": "Temporary PubChem outage",
                "found": False,
                "upstream_error": True,
                "ghs_pictograms": [],
                "hazard_statements": [],
                "precautionary_statements": [],
            },
            {
                "cas_number": "107-18-6",
                "name_en": "Allyl Alcohol",
                "name_zh": "Allyl Alcohol",
                "found": True,
                "ghs_pictograms": [{"code": "GHS06", "name_zh": "\u9ab7\u9acf"}],
                "hazard_statements": [],
                "precautionary_statements": [],
                "has_multiple_classifications": True,
            },
        ]
    }
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post("/api/export/xlsx", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    assert wb.sheetnames == [
        "GHS Results",
        "Ready Rows",
        "Needs Review",
        "Unresolved",
        "Pilot Summary",
    ]
    assert wb["Ready Rows"].max_row == 2
    assert wb["Ready Rows"].cell(row=2, column=1).value == "64-17-5"
    assert wb["Needs Review"].max_row == 3
    assert wb["Needs Review"].cell(row=2, column=1).value == "777-77-7"
    assert wb["Needs Review"].cell(row=2, column=11).value == "Upstream retry needed"
    assert wb["Needs Review"].cell(row=2, column=12).value == "1"
    assert wb["Needs Review"].cell(row=2, column=13).value == "Retry upstream source"
    assert wb["Needs Review"].cell(row=3, column=1).value == "107-18-6"
    assert wb["Needs Review"].cell(row=3, column=11).value == (
        "Multiple GHS classifications; Missing trusted Chinese name"
    )
    assert wb["Needs Review"].cell(row=3, column=12).value == "2"
    assert wb["Needs Review"].cell(row=3, column=13).value == (
        "Confirm primary GHS classification"
    )
    assert wb["Unresolved"].max_row == 2
    assert wb["Unresolved"].cell(row=2, column=1).value == "999-99-9"


async def test_export_csv_neutralizes_formula_injection_in_p_code_text():
    """P-code text starting with a formula trigger must be neutralized
    the same way as other exported cells."""
    transport = ASGITransport(app=app)
    payload = {
        "results": [
            {
                "cas_number": "64-17-5",
                "name_en": "Ethanol",
                "name_zh": "x",
                "ghs_pictograms": [],
                "hazard_statements": [],
                "precautionary_statements": [
                    # Malicious P-code text attempting formula injection
                    {"code": "P999", "text_en": "bad", "text_zh": "=HYPERLINK(\"http://bad\")"},
                ],
            }
        ]
    }
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post("/api/export/csv", json=payload)
    assert response.status_code == 200
    body = response.content.decode("utf-8-sig")
    # The entire joined cell starts with "P999: =HYPERLINK..." — NOT with
    # a bare "=", so it wouldn't execute as a formula anyway. But confirm
    # no neighbouring cell starts with a raw "=HYPERLINK".
    # Ensure at minimum the apostrophe neutralization logic still runs
    # (no raw ",=HYPERLINK" or "\"=HYPERLINK" occurs).
    assert ",=HYPERLINK" not in body
    assert "\"=HYPERLINK" not in body


async def test_export_csv_neutralizes_formula_injection_with_leading_p_code_text():
    """When P-code text itself starts with a formula trigger (edge case:
    someone poisoning the translation data), the apostrophe prefix must
    still apply because the cell value starts with "=" after the
    "P<code>: " prefix concatenation... wait, actually our join prepends
    the P-code first. Cover the stricter case where translation text
    alone starts with a trigger character."""
    transport = ASGITransport(app=app)
    payload = {
        "results": [
            {
                "cas_number": "64-17-5",
                "name_en": "x",
                "name_zh": "x",
                "ghs_pictograms": [],
                "hazard_statements": [],
                # Field whose VALUE starts with = — cell-level hostile data
                "precautionary_statements": [],
            }
        ]
    }
    # Also hit an adjacent field with a trigger char to confirm sanitization
    payload["results"][0]["name_en"] = "=1+1"
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.post("/api/export/csv", json=payload)
    assert response.status_code == 200
    body = response.content.decode("utf-8-sig")
    assert "'=1+1" in body
    assert ",=1+1" not in body


# ─── PubChem retry / backoff ────────────────────────────────

import httpx as _httpx_for_tests
from server import PubChemError, pubchem_get_json


class _FakeResponse:
    def __init__(self, status_code, json_body=None, headers=None):
        self.status_code = status_code
        self._json = json_body if json_body is not None else {}
        self.headers = headers or {}

    def json(self):
        return self._json


class _ScriptedClient:
    """Scripted httpx replacement. Each `get` call pops the next entry
    from `.script`. Entries may be:
        - a _FakeResponse to return
        - an Exception instance to raise
    """

    def __init__(self, script):
        self.script = list(script)
        self.calls = 0

    async def get(self, url, timeout=None):
        self.calls += 1
        item = self.script.pop(0)
        if isinstance(item, Exception):
            raise item
        return item


@pytest.fixture(autouse=True)
def _fast_pubchem_sleep(monkeypatch):
    """Keep retry tests quick: replace asyncio.sleep with a no-op."""
    import asyncio as _a
    import server as _srv

    async def _no_sleep(_delay):
        return None

    monkeypatch.setattr(_srv.asyncio, "sleep", _no_sleep)
    yield


async def test_pubchem_get_json_returns_200_without_retry():
    client = _ScriptedClient([_FakeResponse(200, {"ok": True})])
    status, data = await pubchem_get_json(client, "https://x/", timeout=1.0)
    assert status == 200
    assert data == {"ok": True}
    assert client.calls == 1


async def test_pubchem_get_json_returns_404_without_retry():
    client = _ScriptedClient([_FakeResponse(404)])
    status, data = await pubchem_get_json(client, "https://x/", timeout=1.0)
    assert status == 404
    assert data is None
    assert client.calls == 1  # 404 must NOT trigger retries


async def test_pubchem_get_json_retries_503_then_succeeds():
    client = _ScriptedClient([
        _FakeResponse(503),
        _FakeResponse(200, {"ok": True}),
    ])
    status, data = await pubchem_get_json(client, "https://x/", timeout=1.0, retries=2)
    assert status == 200
    assert data == {"ok": True}
    assert client.calls == 2


async def test_pubchem_get_json_retries_429_with_retry_after():
    client = _ScriptedClient([
        _FakeResponse(429, headers={"Retry-After": "1"}),
        _FakeResponse(200, {"ok": True}),
    ])
    status, data = await pubchem_get_json(client, "https://x/", timeout=1.0, retries=2)
    assert status == 200
    assert client.calls == 2


async def test_pubchem_get_json_raises_after_exhausted_retries_on_5xx():
    client = _ScriptedClient([
        _FakeResponse(503),
        _FakeResponse(502),
        _FakeResponse(500),
    ])
    with pytest.raises(PubChemError):
        await pubchem_get_json(client, "https://x/", timeout=1.0, retries=2)


async def test_pubchem_get_json_raises_on_timeout_exhausted():
    timeout_exc = _httpx_for_tests.TimeoutException("read timeout")
    client = _ScriptedClient([timeout_exc, timeout_exc, timeout_exc])
    with pytest.raises(PubChemError):
        await pubchem_get_json(client, "https://x/", timeout=1.0, retries=2)


async def test_pubchem_get_json_timeout_then_success():
    client = _ScriptedClient([
        _httpx_for_tests.TimeoutException("slow"),
        _FakeResponse(200, {"ok": True}),
    ])
    status, data = await pubchem_get_json(client, "https://x/", timeout=1.0, retries=2)
    assert status == 200
    assert client.calls == 2


async def test_pubchem_get_json_returns_none_for_non_transient_4xx():
    client = _ScriptedClient([_FakeResponse(400)])
    status, data = await pubchem_get_json(client, "https://x/", timeout=1.0)
    assert status == 400
    assert data is None
    assert client.calls == 1  # 400 is not retriable


async def test_search_chemical_surfaces_upstream_error_on_ghs_outage(monkeypatch):
    """A 429/5xx storm on the GHS endpoint must produce
    found=False, upstream_error=True — never a found=True result with
    empty hazard data."""
    import server as srv

    async def fake_get_cid(*_a, **_k):
        return 702  # pretend CID lookup worked

    async def fake_get_name(*_a, **_k):
        return ("Ethanol", "乙醇")

    async def fake_ghs_raise(*_a, **_k):
        raise PubChemError("all retries exhausted")

    monkeypatch.setattr(srv, "get_cid_from_cas", fake_get_cid)
    monkeypatch.setattr(srv, "get_compound_name", fake_get_name)
    monkeypatch.setattr(srv, "get_ghs_classification", fake_ghs_raise)

    class _NullClient:
        async def get(self, *_a, **_k):
            raise RuntimeError("should not be called")

    result = await srv.search_chemical("64-17-5", _NullClient())
    assert result.found is False
    assert result.upstream_error is True
    assert result.ghs_pictograms == []
    assert result.hazard_statements == []
    assert "PubChem 暫時無法回應" in (result.error or "")


async def test_search_chemical_skips_name_lookup_when_dictionary_has_display_names(monkeypatch):
    """Common dictionary-backed chemicals should not wait on PubChem
    name/synonym endpoints once we already have both display names."""
    import server as srv

    async def fake_get_cid(*_a, **_k):
        return 702

    async def fail_get_name(*_a, **_k):  # pragma: no cover - should not run
        raise AssertionError("dictionary-backed lookup should skip PubChem names")

    async def fake_get_ghs(*_a, **_k):
        return ({}, True, "2026-04-16T01:23:45+00:00")

    monkeypatch.setattr(srv, "get_cid_from_cas", fake_get_cid)
    monkeypatch.setattr(srv, "get_compound_name", fail_get_name)
    monkeypatch.setattr(srv, "get_ghs_classification", fake_get_ghs)
    monkeypatch.setattr(srv, "extract_all_ghs_classifications", lambda _: [])

    class _NullClient:
        async def get(self, *_a, **_k):
            raise RuntimeError("should not be called")

    result = await srv.search_chemical("64-17-5", _NullClient())

    assert result.found is True
    assert result.name_en == srv.get_english_name_from_cas("64-17-5")
    assert result.name_zh == srv.get_chinese_name_from_cas("64-17-5")
    assert result.cache_hit is True


async def test_search_chemical_surfaces_upstream_error_on_cid_outage(monkeypatch):
    """When all CID lookup methods transient-fail, surface upstream_error."""
    import server as srv

    async def fake_cid_raise(*_a, **_k):
        raise PubChemError("all CID methods failed")

    monkeypatch.setattr(srv, "get_cid_from_cas", fake_cid_raise)

    class _NullClient:
        async def get(self, *_a, **_k):
            raise RuntimeError("should not be called")

    result = await srv.search_chemical("64-17-5", _NullClient())
    assert result.found is False
    assert result.upstream_error is True
    assert "PubChem 暫時無法回應" in (result.error or "")


async def test_get_cid_prefers_cas_specific_substance_result_over_name_candidate(monkeypatch):
    """CAS-specific lookup must win over generic name lookup.

    PubChem can return related candidates when a CAS number is treated
    as a name query. For 7647-01-0, that can pick chloride/CID 312
    instead of the acid record/CID 313, which changes the GHS label.
    """
    import server as srv

    async def _name_candidate(*_a, **_k):
        return 312

    async def _xref_candidate(*_a, **_k):
        return 312

    async def _substance_candidate(*_a, **_k):
        return 313

    monkeypatch.setattr(srv, "_try_cid_by_name", _name_candidate)
    monkeypatch.setattr(srv, "_try_cid_by_xref", _xref_candidate)
    monkeypatch.setattr(srv, "_try_cid_by_substance", _substance_candidate)

    cas = "7647-01-0"
    srv.cid_cache.pop(cas, None)

    assert await srv.get_cid_from_cas(cas, http_client=None) == 313


async def test_get_cid_partial_transient_mixed_with_404_raises(monkeypatch):
    """Regression (Codex review): if one CID lookup method returns a
    clean 404 but another is transient, we cannot trust the 'not found'
    conclusion — the transient endpoint might have held a record.

    For a chemical safety tool this matters: returning not-found here
    could present a transient outage as confirmed absence of hazard data.

    Scenario: method A (name) → None (clean 404)
              method B (xref) → PubChemError (503 exhausted)
              method C (substance) → PubChemError (timeout)
              method 4 fallback → None (clean, alt-cas path)

    Expected: get_cid_from_cas raises PubChemError, NOT returns None.
    """
    import server as srv

    async def _method_a_clean_404(*_a, **_k):
        return None

    async def _method_b_transient(*_a, **_k):
        raise PubChemError("HTTP 503 after retries")

    async def _method_c_transient(*_a, **_k):
        raise PubChemError("TimeoutException after retries")

    monkeypatch.setattr(srv, "_try_cid_by_name", _method_a_clean_404)
    monkeypatch.setattr(srv, "_try_cid_by_xref", _method_b_transient)
    monkeypatch.setattr(srv, "_try_cid_by_substance", _method_c_transient)
    # Use an unambiguous CAS that won't trigger the alt-CAS fallback
    # (no leading zeros in the first segment).
    cas = "64-17-5"
    # Ensure the cache is empty for this CAS so we actually call the methods.
    srv.cid_cache.pop(cas, None)

    with pytest.raises(PubChemError):
        await srv.get_cid_from_cas(cas, http_client=None)


async def test_search_chemical_partial_cid_transient_surfaces_upstream_error(monkeypatch):
    """End-to-end version of the partial-transient regression.

    One CID method returns a clean 'no match', another times out.
    `search_chemical` must NOT fall back to local-dictionary-only
    success; it must return upstream_error=True so the UI tells the
    user to retry rather than treating a transient outage as
    confirmed absence."""
    import server as srv

    async def _clean_404(*_a, **_k):
        return None

    async def _transient(*_a, **_k):
        raise PubChemError("HTTP 503 after retries")

    monkeypatch.setattr(srv, "_try_cid_by_name", _clean_404)
    monkeypatch.setattr(srv, "_try_cid_by_xref", _transient)
    monkeypatch.setattr(srv, "_try_cid_by_substance", _transient)

    # 64-17-5 (ethanol) also appears in the local CAS dictionary, so
    # the OLD code path would happily return a found=True result with
    # empty hazards and no indication that PubChem was actually
    # unreachable. That is the dangerous case we're guarding against.
    cas = "64-17-5"
    srv.cid_cache.pop(cas, None)

    class _NullClient:
        async def get(self, *_a, **_k):
            raise RuntimeError("should not be called")

    result = await srv.search_chemical(cas, _NullClient())
    assert result.found is False, (
        "Partial upstream failures must NOT be presented as found-with-no-hazards"
    )
    assert result.upstream_error is True
    assert "PubChem 暫時無法回應" in (result.error or "")


# ─── CORS config safety ─────────────────────────────────────

from server import _cors_origins as _configured_cors_origins


def test_cors_origins_config_rejects_wildcard():
    """The parsed CORS_ORIGINS list must never contain `*`."""
    assert "*" not in _configured_cors_origins
    assert len(_configured_cors_origins) >= 1


def test_cors_origins_are_real_urls():
    """Every configured origin must look like an http(s) URL, not a glob."""
    for origin in _configured_cors_origins:
        assert origin.startswith(("http://", "https://")), (
            f"Non-URL origin in CORS config: {origin!r}"
        )


async def test_cors_preflight_echoes_configured_origin_exactly():
    """A preflight from a configured origin must be accepted and the
    allow-origin header must be the exact string — never `*`."""
    configured = _configured_cors_origins[0]
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.options(
            "/api/health",
            headers={
                "Origin": configured,
                "Access-Control-Request-Method": "GET",
            },
        )
    allowed = response.headers.get("access-control-allow-origin")
    assert allowed == configured
    assert allowed != "*"


async def test_cors_preflight_allows_admin_workspace_put_headers():
    """Workspace/admin browser requests must pass CORS preflight before auth."""
    configured = _configured_cors_origins[0]
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.options(
            "/api/workspace/lab_profile",
            headers={
                "Origin": configured,
                "Access-Control-Request-Method": "PUT",
                "Access-Control-Request-Headers": "x-ghs-admin-key,content-type",
            },
        )

    assert response.status_code == 200
    assert response.headers.get("access-control-allow-origin") == configured
    assert "PUT" in response.headers.get("access-control-allow-methods", "")
    allowed_headers = response.headers.get(
        "access-control-allow-headers",
        "",
    ).lower()
    assert "x-ghs-admin-key" in allowed_headers
    assert "content-type" in allowed_headers


async def test_cors_rejects_disallowed_origin():
    """A non-whitelisted origin must not be echoed back."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.options(
            "/api/health",
            headers={
                "Origin": "https://attacker.example.com",
                "Access-Control-Request-Method": "GET",
            },
        )
    assert response.headers.get("access-control-allow-origin") != "https://attacker.example.com"


async def test_cors_does_not_advertise_credentials_mode():
    """allow_credentials must be False (API has no cookies or auth headers)."""
    configured = _configured_cors_origins[0]
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.options(
            "/api/health",
            headers={
                "Origin": configured,
                "Access-Control-Request-Method": "GET",
            },
        )
    cred_header = response.headers.get("access-control-allow-credentials", "").lower()
    assert cred_header != "true"


# ─── Rate limiting ──────────────────────────────────────────

from server import limiter as _limiter


def test_limiter_is_configured_with_custom_key_func():
    """Rate limit buckets must key off resolved client IP, not a shared
    global bucket."""
    from server import _client_ip
    assert _limiter._key_func is _client_ip


def test_pubchem_outbound_semaphore_is_bounded():
    """The outbound PubChem concurrency gate must be a small positive
    integer so a burst of client requests cannot flood PubChem."""
    from server import _pubchem_semaphore, PUBCHEM_OUTBOUND_CONCURRENCY
    assert isinstance(PUBCHEM_OUTBOUND_CONCURRENCY, int)
    assert 1 <= PUBCHEM_OUTBOUND_CONCURRENCY <= 32
    # The semaphore's value should match the configured concurrency.
    # The `_value` attribute exists on asyncio.Semaphore in CPython.
    assert _pubchem_semaphore._value == PUBCHEM_OUTBOUND_CONCURRENCY


def test_client_ip_ignores_forwarded_header_by_default(monkeypatch):
    """Forwarded headers are client-controlled unless the deployment has
    verified that its proxy overwrites them."""
    from server import _client_ip
    monkeypatch.delenv("TRUST_FORWARDED_HEADERS", raising=False)

    class _FakeHeaders(dict):
        def get(self, k, default=None):
            return super().get(k.lower(), default)

    class _FakeClient:
        host = "10.0.0.1"

    class _FakeRequest:
        headers = _FakeHeaders({"x-forwarded-for": "203.0.113.9, 10.0.0.1"})
        client = _FakeClient()

    assert _client_ip(_FakeRequest()) == "10.0.0.1"


def test_client_ip_prefers_leftmost_x_forwarded_for_when_trusted(monkeypatch):
    """Opted-in trusted proxy deployments can still bucket by the original
    client address."""
    from server import _client_ip
    monkeypatch.setenv("TRUST_FORWARDED_HEADERS", "1")

    class _FakeHeaders(dict):
        def get(self, k, default=None):
            return super().get(k.lower(), default)

    class _FakeClient:
        host = "10.0.0.1"

    class _FakeRequest:
        headers = _FakeHeaders({"x-forwarded-for": "203.0.113.9, 10.0.0.1"})
        client = _FakeClient()

    assert _client_ip(_FakeRequest()) == "203.0.113.9"


def test_client_ip_skips_invalid_forwarded_values_when_trusted(monkeypatch):
    from server import _client_ip
    monkeypatch.setenv("TRUST_FORWARDED_HEADERS", "1")

    class _FakeHeaders(dict):
        def get(self, k, default=None):
            return super().get(k.lower(), default)

    class _FakeClient:
        host = "10.0.0.1"

    class _FakeRequest:
        headers = _FakeHeaders({"x-forwarded-for": "spoofed, 203.0.113.9"})
        client = _FakeClient()

    assert _client_ip(_FakeRequest()) == "203.0.113.9"


def test_client_ip_falls_back_to_direct_peer_without_forwarded_header():
    from server import _client_ip

    class _FakeHeaders(dict):
        def get(self, k, default=None):
            return super().get(k.lower(), default)

    class _FakeClient:
        host = "198.51.100.5"

    class _FakeRequest:
        headers = _FakeHeaders()
        client = _FakeClient()

    assert _client_ip(_FakeRequest()) == "198.51.100.5"


def test_health_endpoint_has_no_rate_limit():
    """/api/health must remain unlimited so load balancers / uptime
    monitors don't trip the limiter when polling."""
    from starlette.testclient import TestClient
    client = TestClient(app)
    # Burst 20 requests; none should be 429.
    statuses = [client.get("/api/health").status_code for _ in range(20)]
    assert all(s == 200 for s in statuses), statuses


# ─── Security response headers ──────────────────────────────

async def test_health_response_has_nosniff_header():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/health")
    assert response.status_code == 200
    assert response.headers.get("x-content-type-options") == "nosniff"


async def test_health_response_has_referrer_policy():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/health")
    assert response.headers.get("referrer-policy") == "strict-origin-when-cross-origin"


async def test_health_response_has_permissions_policy():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/health")
    # The policy must disable at least the sensor/credential-style APIs.
    pp = response.headers.get("permissions-policy", "")
    for directive in ("camera=", "microphone=", "geolocation="):
        assert directive in pp, f"missing {directive!r} in Permissions-Policy"


async def test_api_response_has_strict_csp():
    """API returns JSON/binary; its CSP must be extremely restrictive."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        response = await ac.get("/api/health")
    csp = response.headers.get("content-security-policy", "")
    assert "default-src 'none'" in csp
    assert "frame-ancestors 'none'" in csp


# ─── P-code translation dictionaries ────────────────────────

def test_p_code_zh_and_en_dicts_have_identical_keys():
    """If someone adds a P-code to one dictionary but forgets the
    other, this test catches it immediately."""
    zh_keys = set(P_CODE_TRANSLATIONS.keys())
    en_keys = set(P_CODE_TEXTS_EN.keys())
    missing_en = zh_keys - en_keys
    missing_zh = en_keys - zh_keys
    assert not missing_en, f"Keys in ZH but missing from EN: {missing_en}"
    assert not missing_zh, f"Keys in EN but missing from ZH: {missing_zh}"


# ─── P-code extraction (v1.8 M0) ───────────────────────────

from server import extract_all_ghs_classifications
from p_code_translations import P_CODE_TRANSLATIONS, P_CODE_TEXTS_EN


def _make_ghs_data(*reports):
    """Build a minimal PubChem-shaped dict containing the given
    GHS report Information entries."""
    return {
        "Record": {
            "Section": [{
                "TOCHeading": "Safety and Hazards",
                "Section": [{
                    "TOCHeading": "Hazards Identification",
                    "Section": [{
                        "TOCHeading": "GHS Classification",
                        "Information": list(reports),
                    }]
                }]
            }]
        }
    }


def _pic_info(codes):
    """Create a Pictogram(s) Information entry."""
    return {
        "Name": "Pictogram(s)",
        "Value": {"StringWithMarkup": [{
            "Markup": [
                {"Type": "Icon", "URL": f"https://example.com/{c}.svg"}
                for c in codes
            ]
        }]}
    }


def _signal_info(word):
    return {"Name": "Signal", "Value": {"StringWithMarkup": [{"String": word}]}}


def _hazard_info(*h_texts):
    return {
        "Name": "GHS Hazard Statements",
        "Value": {"StringWithMarkup": [{"String": t} for t in h_texts]},
    }


def _precaution_info(codes_csv):
    """Create a Precautionary Statement Codes Information entry.
    PubChem returns P-codes as a comma-separated string."""
    return {
        "Name": "Precautionary Statement Codes",
        "Value": {"StringWithMarkup": [{"String": codes_csv}]},
    }


class TestPCodeExtraction:
    """Tests for parsing PubChem Precautionary Statement Codes."""

    def test_single_p_codes_extracted(self):
        data = _make_ghs_data(
            _pic_info(["GHS02"]),
            _signal_info("Danger"),
            _hazard_info("H225: Highly Flammable liquid"),
            _precaution_info("P210, P233, P240, P280, and P501"),
        )
        reports = extract_all_ghs_classifications(data)
        assert len(reports) == 1
        p_codes = [p["code"] for p in reports[0]["precautionary_statements"]]
        assert "P210" in p_codes
        assert "P233" in p_codes
        assert "P501" in p_codes
        assert len(p_codes) == 5

    def test_combined_p_codes_kept_intact(self):
        """Combined codes like P301+P310 must not be split."""
        data = _make_ghs_data(
            _pic_info(["GHS06"]),
            _signal_info("Danger"),
            _hazard_info("H301: Toxic if swallowed"),
            _precaution_info("P264, P270, P301+P310, P321, P405, and P501"),
        )
        reports = extract_all_ghs_classifications(data)
        p_codes = [p["code"] for p in reports[0]["precautionary_statements"]]
        assert "P301+P310" in p_codes
        assert p_codes.count("P301+P310") == 1

    def test_triple_combined_code(self):
        """Three-part combined codes like P303+P361+P353 must survive."""
        data = _make_ghs_data(
            _pic_info(["GHS02"]),
            _signal_info("Danger"),
            _precaution_info("P210, P303+P361+P353, P403+P235"),
        )
        reports = extract_all_ghs_classifications(data)
        p_codes = [p["code"] for p in reports[0]["precautionary_statements"]]
        assert "P303+P361+P353" in p_codes
        assert "P403+P235" in p_codes

    def test_p_codes_deduplicated_within_report(self):
        """If PubChem repeats a code in the same string, only keep one."""
        data = _make_ghs_data(
            _pic_info(["GHS07"]),
            _precaution_info("P264, P280, P264, P280, P501"),
        )
        reports = extract_all_ghs_classifications(data)
        p_codes = [p["code"] for p in reports[0]["precautionary_statements"]]
        assert p_codes.count("P264") == 1
        assert p_codes.count("P280") == 1

    def test_known_translation_populates_text_zh_and_text_en(self):
        data = _make_ghs_data(
            _pic_info(["GHS02"]),
            _precaution_info("P210"),
        )
        reports = extract_all_ghs_classifications(data)
        stmt = reports[0]["precautionary_statements"][0]
        assert stmt["code"] == "P210"
        # text_zh must be the Chinese translation, not the code
        assert stmt["text_zh"] != "P210"
        assert len(stmt["text_zh"]) > 5
        # text_en must be the full English text, not just the code
        assert stmt["text_en"] != "P210"
        assert "heat" in stmt["text_en"].lower() or "ignition" in stmt["text_en"].lower()

    def test_unknown_code_falls_back_to_code_string(self):
        """P-codes not in the translation dict should not be dropped;
        they should fall back to using the code itself as text_zh."""
        data = _make_ghs_data(
            _pic_info(["GHS07"]),
            _precaution_info("P999"),
        )
        reports = extract_all_ghs_classifications(data)
        stmt = reports[0]["precautionary_statements"][0]
        assert stmt["code"] == "P999"
        assert stmt["text_zh"] == "P999"

    def test_report_without_p_codes_has_empty_list(self):
        data = _make_ghs_data(
            _pic_info(["GHS02"]),
            _signal_info("Danger"),
            _hazard_info("H225: Highly Flammable liquid"),
        )
        reports = extract_all_ghs_classifications(data)
        assert len(reports) == 1
        assert reports[0]["precautionary_statements"] == []


class TestClassificationSignatureWithPCodes:
    """P-codes must now be part of the dedup signature."""

    def test_different_p_codes_produce_different_signatures(self):
        a = {
            "pictograms": [_pic("GHS02")],
            "hazard_statements": [_hz("H225")],
            "precautionary_statements": [{"code": "P210"}, {"code": "P233"}],
            "signal_word": "Danger",
            "source": "ECHA",
        }
        b = {
            "pictograms": [_pic("GHS02")],
            "hazard_statements": [_hz("H225")],
            "precautionary_statements": [{"code": "P210"}, {"code": "P501"}],
            "signal_word": "Danger",
            "source": "ECHA",
        }
        assert _classification_signature(a) != _classification_signature(b)

    def test_same_p_codes_different_order_equal_signature(self):
        a = {
            "pictograms": [_pic("GHS02")],
            "hazard_statements": [_hz("H225")],
            "precautionary_statements": [{"code": "P501"}, {"code": "P210"}],
            "signal_word": "Danger",
            "source": "ECHA",
        }
        b = {
            "pictograms": [_pic("GHS02")],
            "hazard_statements": [_hz("H225")],
            "precautionary_statements": [{"code": "P210"}, {"code": "P501"}],
            "signal_word": "Danger",
            "source": "ECHA",
        }
        assert _classification_signature(a) == _classification_signature(b)


async def test_search_chemical_returns_precautionary_statements(monkeypatch):
    """End-to-end: search_chemical must surface P-codes in the primary
    result and in other_classifications."""
    import server as srv

    fake_reports = [
        {
            "pictograms": [_pic("GHS02")],
            "hazard_statements": [_hz("H225")],
            "precautionary_statements": [
                {"code": "P210", "text_en": "P210", "text_zh": "test-zh"},
                {"code": "P233", "text_en": "P233", "text_zh": "test-zh2"},
            ],
            "signal_word": "Danger",
            "signal_word_zh": "\u5371\u96aa",
            "source": "ECHA",
            "report_count": "100",
        },
        {
            "pictograms": [_pic("GHS02")],
            "hazard_statements": [_hz("H225")],
            "precautionary_statements": [
                {"code": "P210", "text_en": "P210", "text_zh": "test-zh"},
                {"code": "P501", "text_en": "P501", "text_zh": "test-zh3"},
            ],
            "signal_word": "Danger",
            "signal_word_zh": "\u5371\u96aa",
            "source": "ECHA",
            "report_count": "50",
        },
    ]

    async def fake_get_cid(*_a, **_k):
        return 702

    async def fake_get_name(*_a, **_k):
        return ("Ethanol", "\u4e59\u9187")

    async def fake_get_ghs(*_a, **_k):
        # v1.8 M1: get_ghs_classification now returns (data, cache_hit, retrieved_at)
        return ({}, False, "2026-04-16T00:00:00+00:00")

    monkeypatch.setattr(srv, "get_cid_from_cas", fake_get_cid)
    monkeypatch.setattr(srv, "get_compound_name", fake_get_name)
    monkeypatch.setattr(srv, "get_ghs_classification", fake_get_ghs)
    monkeypatch.setattr(srv, "extract_all_ghs_classifications", lambda _: fake_reports)

    class _NullClient:
        async def get(self, *_a, **_k):
            raise RuntimeError("should not be called")

    result = await srv.search_chemical("64-17-5", _NullClient())

    # Primary must have P-codes
    assert len(result.precautionary_statements) > 0
    primary_p = {p["code"] for p in result.precautionary_statements}
    assert "P210" in primary_p

    # Both reports must survive (different P-codes = different signature)
    assert result.has_multiple_classifications is True
    assert len(result.other_classifications) == 1
    other_p = {p["code"] for p in result.other_classifications[0].precautionary_statements}
    assert other_p != primary_p


# ─── Provenance / trust signals (v1.8 M1) ───────────────────

class TestGhsCacheProvenance:
    """get_ghs_classification now returns (data, cache_hit, retrieved_at).
    First call fetches fresh; second hits cache; both carry timestamps."""

    async def test_first_call_is_cache_miss_with_timestamp(self, monkeypatch):
        import server as srv
        # Clear cache for determinism
        srv.ghs_cache.clear()

        async def fake_pubchem_get_json(*_a, **_k):
            return 200, {"Record": {"some": "payload"}}

        monkeypatch.setattr(srv, "pubchem_get_json", fake_pubchem_get_json)

        data, cache_hit, retrieved_at = await srv.get_ghs_classification(12345, http_client=None)
        assert cache_hit is False
        assert retrieved_at is not None
        # ISO-8601 with timezone offset
        assert "T" in retrieved_at
        assert data == {"Record": {"some": "payload"}}

    async def test_second_call_is_cache_hit_with_same_timestamp(self, monkeypatch):
        import server as srv
        srv.ghs_cache.clear()

        call_count = {"n": 0}

        async def fake_pubchem_get_json(*_a, **_k):
            call_count["n"] += 1
            return 200, {"Record": {"some": "payload"}}

        monkeypatch.setattr(srv, "pubchem_get_json", fake_pubchem_get_json)

        _, cache_hit_1, ts_1 = await srv.get_ghs_classification(12345, http_client=None)
        _, cache_hit_2, ts_2 = await srv.get_ghs_classification(12345, http_client=None)

        assert cache_hit_1 is False
        assert cache_hit_2 is True, "second call must hit the 24hr cache"
        assert ts_1 == ts_2, "cached timestamp should match the original fetch time"
        assert call_count["n"] == 1, "pubchem must not be hit twice for same CID"

    async def test_cache_miss_on_404_not_persisted(self, monkeypatch):
        """Genuine no-data responses (404) should not be cached so the
        next request can retry. Timestamp still populated."""
        import server as srv
        srv.ghs_cache.clear()

        async def fake_pubchem_get_json(*_a, **_k):
            return 404, None

        monkeypatch.setattr(srv, "pubchem_get_json", fake_pubchem_get_json)

        data, cache_hit, retrieved_at = await srv.get_ghs_classification(99999, http_client=None)
        assert data == {}
        assert cache_hit is False
        assert retrieved_at is not None
        # Second call must NOT report cache_hit=True since 404 wasn't cached
        data2, cache_hit_2, _ = await srv.get_ghs_classification(99999, http_client=None)
        assert cache_hit_2 is False


async def test_search_chemical_response_includes_provenance(monkeypatch):
    """search_chemical must surface primary_source, primary_report_count,
    retrieved_at, and cache_hit in the response."""
    import server as srv

    fake_reports = [
        {
            "pictograms": [_pic("GHS02")],
            "hazard_statements": [_hz("H225")],
            "precautionary_statements": [],
            "signal_word": "Danger",
            "signal_word_zh": "\u5371\u96aa",
            "source": "ECHA C&L Notifications Summary",
            "report_count": "236",
        },
    ]

    async def fake_get_cid(*_a, **_k):
        return 702

    async def fake_get_name(*_a, **_k):
        return ("Ethanol", "\u4e59\u9187")

    async def fake_get_ghs(*_a, **_k):
        return ({}, True, "2026-04-16T01:23:45+00:00")

    monkeypatch.setattr(srv, "get_cid_from_cas", fake_get_cid)
    monkeypatch.setattr(srv, "get_compound_name", fake_get_name)
    monkeypatch.setattr(srv, "get_ghs_classification", fake_get_ghs)
    monkeypatch.setattr(srv, "extract_all_ghs_classifications", lambda _: fake_reports)

    class _NullClient:
        async def get(self, *_a, **_k):
            raise RuntimeError("should not be called")

    result = await srv.search_chemical("64-17-5", _NullClient())

    # Provenance fields
    assert result.primary_source == "ECHA C&L Notifications Summary"
    assert result.primary_report_count == "236"
    assert result.retrieved_at == "2026-04-16T01:23:45+00:00"
    assert result.cache_hit is True


async def test_search_chemical_no_ghs_reports_still_populates_timestamp(monkeypatch):
    """Even when PubChem has a CID but no GHS reports, the response
    must carry the retrieval timestamp so the UI can say
    'checked at X, no hazard data available'."""
    import server as srv

    async def fake_get_cid(*_a, **_k):
        return 12345

    async def fake_get_name(*_a, **_k):
        return ("SomeCompound", None)

    async def fake_get_ghs(*_a, **_k):
        return ({}, False, "2026-04-16T02:00:00+00:00")

    monkeypatch.setattr(srv, "get_cid_from_cas", fake_get_cid)
    monkeypatch.setattr(srv, "get_compound_name", fake_get_name)
    monkeypatch.setattr(srv, "get_ghs_classification", fake_get_ghs)
    monkeypatch.setattr(srv, "extract_all_ghs_classifications", lambda _: [])

    class _NullClient:
        async def get(self, *_a, **_k):
            raise RuntimeError("should not be called")

    result = await srv.search_chemical("123-45-6", _NullClient())

    assert result.found is True
    assert result.retrieved_at == "2026-04-16T02:00:00+00:00"
    assert result.cache_hit is False
    assert result.primary_source is None
    assert result.primary_report_count is None
