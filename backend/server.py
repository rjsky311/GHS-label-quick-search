from collections import Counter, deque
from contextlib import asynccontextmanager
from fastapi import FastAPI, APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware
import os
import secrets
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import httpx
import asyncio
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import csv
import random
import re
from cachetools import TTLCache
from pilot_store import APPROVED_ALIAS_STATUS, PilotStore, infer_locale, normalize_compact_text

# Import expanded chemical dictionaries (1707 CAS entries, 1816 English entries)
# + alias dictionaries for common/colloquial chemical names
from chemical_dict import (
    CAS_TO_ZH, CAS_TO_EN, CHEMICAL_NAMES_ZH_EXPANDED,
    EN_TO_CAS, ZH_TO_CAS, ALIASES_ZH, ALIASES_EN,
)
from p_code_translations import P_CODE_TRANSLATIONS, P_CODE_TEXTS_EN

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')


APP_VERSION = "1.10.0"
WORKSPACE_DOC_TYPES = {
    "lab_profile",
    "print_custom_label_fields",
    "print_recents",
    "print_templates",
    "prepared_recents",
    "prepared_presets",
}
PILOT_STORE_PATH = Path(os.environ.get("PILOT_STORE_PATH") or (ROOT_DIR / "data" / "pilot.db"))
pilot_store = PilotStore(PILOT_STORE_PATH)
ADMIN_API_TOKEN = (os.environ.get("ADMIN_API_TOKEN") or "").strip()
CAPTURE_DICTIONARY_MISSES = (
    (os.environ.get("CAPTURE_DICTIONARY_MISSES") or "").strip().lower()
    in {"1", "true", "yes", "on"}
)

# Shared httpx client (initialized in lifespan)
shared_http_client: Optional[httpx.AsyncClient] = None

# In-memory caches (TTL = 24 hours, max 5000 entries each)
cid_cache: TTLCache = TTLCache(maxsize=5000, ttl=86400)
ghs_cache: TTLCache = TTLCache(maxsize=5000, ttl=86400)
ops_counters: Counter = Counter()
ops_recent_events = deque(maxlen=50)
OPS_STALE_THRESHOLD_HOURS = float(os.environ.get("OPS_STALE_THRESHOLD_HOURS", "12"))

# ─── Outbound PubChem concurrency gate ──────────────────────
#
# Limits how many concurrent PubChem requests this process can have in
# flight regardless of how many client requests come in. Without this
# gate, a single burst of batch searches (100 CAS numbers × 3-4 PubChem
# calls each) could blow past PubChem's published rate limits and get
# the whole deploy temporarily blocked. 8 is a conservative default
# that still benefits from the shared httpx client's keep-alive pool.
PUBCHEM_OUTBOUND_CONCURRENCY = int(os.environ.get("PUBCHEM_CONCURRENCY", "8"))
_pubchem_semaphore = asyncio.Semaphore(PUBCHEM_OUTBOUND_CONCURRENCY)


def _record_ops_counter(key: str, amount: int = 1) -> None:
    ops_counters[key] += amount


def _record_ops_event(event_type: str, **payload: Any) -> None:
    ops_recent_events.append(
        {
            "ts": datetime.now(timezone.utc).isoformat(),
            "type": event_type,
            **payload,
        }
    )


def _parse_iso_ts(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


def _cache_age_hours(retrieved_at: Optional[str]) -> Optional[float]:
    parsed = _parse_iso_ts(retrieved_at)
    if parsed is None:
        return None
    return max(0.0, (datetime.now(timezone.utc) - parsed).total_seconds() / 3600)


def _cache_age_bucket(age_hours: float) -> str:
    if age_hours < 1:
        return "<1h"
    if age_hours < 6:
        return "1-6h"
    if age_hours < 12:
        return "6-12h"
    if age_hours < 24:
        return "12-24h"
    return "24h+"


def _observe_ghs_cache_hit(cid: int, retrieved_at: Optional[str]) -> None:
    _record_ops_counter("cache.ghs.hit")
    age_hours = _cache_age_hours(retrieved_at)
    if age_hours is None:
        return
    _record_ops_counter(f"cache.ghs.age_bucket.{_cache_age_bucket(age_hours)}")
    if age_hours >= OPS_STALE_THRESHOLD_HOURS:
        _record_ops_counter("cache.ghs.stale_hit")
        _record_ops_event(
            "cache_stale_hit",
            cid=cid,
            age_hours=round(age_hours, 2),
            threshold_hours=OPS_STALE_THRESHOLD_HOURS,
        )


def _record_upstream_failure(kind: str, url: str, attempt: int, **payload: Any) -> None:
    _record_ops_counter("upstream.total")
    _record_ops_counter(f"upstream.{kind}")
    _record_ops_event(
        "upstream_error",
        kind=kind,
        attempt=attempt,
        url=url,
        **payload,
    )


def _require_admin(request: Request) -> None:
    if not ADMIN_API_TOKEN:
        raise HTTPException(
            status_code=503,
            detail="Admin API is not configured on this server.",
        )

    supplied_token = (request.headers.get("x-ghs-admin-key") or "").strip()
    if not supplied_token:
        raise HTTPException(status_code=401, detail="Admin key required.")

    if not secrets.compare_digest(supplied_token, ADMIN_API_TOKEN):
        raise HTTPException(status_code=403, detail="Invalid admin key.")


def _ensure_workspace_doc_type(doc_type: str) -> str:
    if doc_type not in WORKSPACE_DOC_TYPES:
        raise HTTPException(status_code=404, detail=f"Unsupported workspace document: {doc_type}")
    return doc_type


def _is_cas_like_query(query: str) -> bool:
    return bool(re.match(r"^[\d-]+$", query.strip()))


def _manual_name_pairs(locale: str) -> list[tuple[str, str]]:
    pairs = []
    for entry in pilot_store.list_manual_entries():
        if locale == "zh":
            name = (entry.get("name_zh") or "").strip()
        else:
            name = (entry.get("name_en") or "").strip().lower()
        if name:
            pairs.append((name, entry["cas_number"]))
    return pairs


def _approved_alias_pairs(locale: str) -> list[tuple[str, str]]:
    pairs = []
    for alias in pilot_store.list_aliases(status=APPROVED_ALIAS_STATUS, locale=locale):
        key = (alias.get("alias_text") or "").strip()
        if not key:
            continue
        pairs.append((key.lower() if locale == "en" else key, alias["cas_number"]))
    return pairs


def _combined_lookup_map(locale: str) -> Dict[str, str]:
    if locale == "zh":
        combined = dict(ZH_TO_CAS)
    else:
        combined = dict(EN_TO_CAS)

    for key, cas in _manual_name_pairs(locale):
        combined[key] = cas

    for key, cas in _approved_alias_pairs(locale):
        if key not in combined:
            combined[key] = cas

    return combined


def _compact_exact_match(query: str, locale: str) -> Optional[str]:
    compact_query = normalize_compact_text(query, locale=locale)
    if not compact_query:
        return None

    matches = set()
    for key, cas in _combined_lookup_map(locale).items():
        if normalize_compact_text(key, locale=locale) == compact_query:
            matches.add(cas)
    if len(matches) == 1:
        return next(iter(matches))
    return None


def _build_reference_links(
    cas_number: Optional[str],
    cid: Optional[int],
    name_en: Optional[str] = None,
) -> List[Dict[str, Any]]:
    links: list[dict[str, Any]] = []
    seen_urls: set[str] = set()

    if cid:
        links.append(
            {
                "label": "PubChem Safety & Hazards",
                "url": f"https://pubchem.ncbi.nlm.nih.gov/compound/{cid}#section=Safety-and-Hazards",
                "link_type": "sds",
                "source": "pubchem",
                "priority": 10,
            }
        )
        links.append(
            {
                "label": "PubChem Compound Overview",
                "url": f"https://pubchem.ncbi.nlm.nih.gov/compound/{cid}",
                "link_type": "reference",
                "source": "pubchem",
                "priority": 40,
            }
        )
    if cas_number:
        links.append(
            {
                "label": "ECHA Substance Search",
                "url": f"https://chem.echa.europa.eu/substance-search?searchText={cas_number}",
                "link_type": "regulatory",
                "source": "echa",
                "priority": 20,
            }
        )
    if cas_number or name_en:
        links.append(
            {
                "label": "NIOSH Pocket Guide",
                "url": "https://www.cdc.gov/niosh/npg/default.html",
                "link_type": "occupational",
                "source": "niosh",
                "priority": 60,
            }
        )

    if cas_number:
        for link in pilot_store.list_reference_links(cas_number):
            links.append(
                {
                    "label": link["label"],
                    "url": link["url"],
                    "link_type": link["linkType"],
                    "source": link["source"],
                    "priority": link["priority"],
                }
            )

    deduped = []
    for link in sorted(links, key=lambda item: (item["priority"], item["label"])):
        if link["url"] in seen_urls:
            continue
        seen_urls.add(link["url"])
        deduped.append(link)
    return deduped


def _record_dictionary_miss(query: str, query_kind: str, endpoint: str, *, context: Optional[Dict[str, Any]] = None) -> None:
    if not CAPTURE_DICTIONARY_MISSES:
        return

    try:
        pilot_store.record_miss_query(
            query,
            query_kind,
            endpoint,
            context=context or {},
        )
    except Exception as exc:
        logger.warning("Failed to persist miss query '%s': %s", query, exc)


# ─── Client IP resolution for rate limiting ─────────────────
#
# The service sits behind Zeabur's proxy, so `request.client.host`
# always resolves to the proxy. Without taking X-Forwarded-For into
# account, every user would share one rate-limit bucket.
#
# Trust the LEFTMOST (original client) IP in X-Forwarded-For when
# present, and fall back to the direct peer address otherwise.
#
# Deployment caveat (documented here and in README): this assumes the
# API is only reachable through a trusted reverse proxy. If the API
# is ever exposed directly to the public, a malicious client could
# forge X-Forwarded-For. In that case set TRUST_FORWARDED_HEADERS=0.
def _client_ip(request: "Request") -> str:
    if os.environ.get("TRUST_FORWARDED_HEADERS", "1") == "1":
        forwarded = request.headers.get("x-forwarded-for")
        if forwarded:
            first = forwarded.split(",")[0].strip()
            if first:
                return first
    if request.client:
        return request.client.host
    return "unknown"


# slowapi limiter. In-memory bucket — correct for a single-worker
# deployment. Scaling to multiple workers/instances requires swapping
# in a Redis-backed storage URI or pushing rate-limiting out to the
# edge (Zeabur router, Cloudflare, etc.).
limiter = Limiter(
    key_func=_client_ip,
    default_limits=[],  # no implicit global limit; per-route only
    # headers_enabled=False so slowapi's decorator does not try to
    # pluck a Response from each endpoint signature. Rate-limit
    # signalling happens through the 429 status + exception handler.
    headers_enabled=False,
)

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifespan: startup and shutdown events."""
    global shared_http_client
    pilot_store.connect()
    shared_http_client = httpx.AsyncClient(
        timeout=30.0,
        limits=httpx.Limits(max_connections=20, max_keepalive_connections=10),
    )
    yield
    # Shutdown
    await shared_http_client.aclose()
    pilot_store.close()

# Create the main app with lifespan
app = FastAPI(title="GHS Label Quick Search API", lifespan=lifespan)

# Register the rate limiter. The exception handler translates
# RateLimitExceeded into a 429 response so clients see a clear signal
# rather than a generic 500.
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
app.add_middleware(SlowAPIMiddleware)

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# GHS Pictogram mapping
GHS_PICTOGRAMS = {
    "GHS01": {"name": "Explosive", "name_zh": "爆炸物", "icon": "💥", "image": "https://pubchem.ncbi.nlm.nih.gov/images/ghs/GHS01.svg"},
    "GHS02": {"name": "Flammable", "name_zh": "易燃物", "icon": "🔥", "image": "https://pubchem.ncbi.nlm.nih.gov/images/ghs/GHS02.svg"},
    "GHS03": {"name": "Oxidizer", "name_zh": "氧化劑", "icon": "⭕", "image": "https://pubchem.ncbi.nlm.nih.gov/images/ghs/GHS03.svg"},
    "GHS04": {"name": "Compressed Gas", "name_zh": "壓縮氣體", "icon": "🫧", "image": "https://pubchem.ncbi.nlm.nih.gov/images/ghs/GHS04.svg"},
    "GHS05": {"name": "Corrosive", "name_zh": "腐蝕性", "icon": "🧪", "image": "https://pubchem.ncbi.nlm.nih.gov/images/ghs/GHS05.svg"},
    "GHS06": {"name": "Toxic", "name_zh": "劇毒", "icon": "💀", "image": "https://pubchem.ncbi.nlm.nih.gov/images/ghs/GHS06.svg"},
    "GHS07": {"name": "Irritant", "name_zh": "刺激性/有害", "icon": "⚠️", "image": "https://pubchem.ncbi.nlm.nih.gov/images/ghs/GHS07.svg"},
    "GHS08": {"name": "Health Hazard", "name_zh": "健康危害", "icon": "🫁", "image": "https://pubchem.ncbi.nlm.nih.gov/images/ghs/GHS08.svg"},
    "GHS09": {"name": "Environmental Hazard", "name_zh": "環境危害", "icon": "🐟", "image": "https://pubchem.ncbi.nlm.nih.gov/images/ghs/GHS09.svg"},
}

# H-code Chinese translations
H_CODE_TRANSLATIONS = {
    "H200": "不穩定爆炸物",
    "H201": "爆炸物；整體爆炸危險",
    "H202": "爆炸物；嚴重拋射危險",
    "H203": "爆炸物；火災、爆炸或拋射危險",
    "H204": "火災或拋射危險",
    "H205": "遇火可能整體爆炸",
    "H220": "極易燃氣體",
    "H221": "易燃氣體",
    "H222": "極易燃氣溶膠",
    "H223": "易燃氣溶膠",
    "H224": "極易燃液體和蒸氣",
    "H225": "高度易燃液體和蒸氣",
    "H226": "易燃液體和蒸氣",
    "H227": "可燃液體",
    "H228": "易燃固體",
    "H229": "壓力容器：遇熱可能爆裂",
    "H230": "可能以爆炸方式反應，即使沒有空氣",
    "H231": "在高壓/高溫下可能以爆炸方式反應，即使沒有空氣",
    "H240": "遇熱可能爆炸",
    "H241": "遇熱可能起火或爆炸",
    "H242": "遇熱可能起火",
    "H250": "暴露在空氣中會自燃",
    "H251": "自熱；可能起火",
    "H252": "大量堆積時自熱；可能起火",
    "H260": "遇水放出易燃氣體，可能自燃",
    "H261": "遇水放出易燃氣體",
    "H270": "可能導致或加劇燃燒；氧化劑",
    "H271": "可能引起燃燒或爆炸；強氧化劑",
    "H272": "可能加劇燃燒；氧化劑",
    "H280": "內含高壓氣體；遇熱可能爆炸",
    "H281": "內含冷凍氣體；可能造成低溫灼傷",
    "H290": "可能腐蝕金屬",
    "H300": "吞食致命",
    "H301": "吞食有毒",
    "H302": "吞食有害",
    "H303": "吞食可能有害",
    "H304": "吞食並進入呼吸道可能致命",
    "H305": "吞食並進入呼吸道可能有害",
    "H310": "皮膚接觸致命",
    "H311": "皮膚接觸有毒",
    "H312": "皮膚接觸有害",
    "H313": "皮膚接觸可能有害",
    "H314": "造成嚴重皮膚灼傷和眼睛損傷",
    "H315": "造成皮膚刺激",
    "H316": "造成輕微皮膚刺激",
    "H317": "可能造成皮膚過敏反應",
    "H318": "造成嚴重眼睛損傷",
    "H319": "造成嚴重眼睛刺激",
    "H320": "造成眼睛刺激",
    "H330": "吸入致命",
    "H331": "吸入有毒",
    "H332": "吸入有害",
    "H333": "吸入可能有害",
    "H334": "吸入可能導致過敏或哮喘症狀或呼吸困難",
    "H335": "可能造成呼吸道刺激",
    "H336": "可能造成昏睡或頭暈",
    "H340": "可能導致遺傳性缺陷",
    "H341": "懷疑會導致遺傳性缺陷",
    "H350": "可能致癌",
    "H351": "懷疑會致癌",
    "H360": "可能損害生育能力或胎兒",
    "H361": "懷疑會損害生育能力或胎兒",
    "H362": "可能對哺乳兒童造成傷害",
    "H370": "會對器官造成損害",
    "H371": "可能會對器官造成損害",
    "H372": "長期或反覆暴露會對器官造成損害",
    "H373": "長期或反覆暴露可能會對器官造成損害",
    "H400": "對水生生物毒性非常大",
    "H401": "對水生生物有毒",
    "H402": "對水生生物有害",
    "H410": "對水生生物毒性非常大並具有長期持續影響",
    "H411": "對水生生物有毒並具有長期持續影響",
    "H412": "對水生生物有害並具有長期持續影響",
    "H413": "可能對水生生物造成長期持續有害影響",
    "H420": "破壞高層大氣中的臭氧，危害公眾健康和環境",
}

# Pre-computed cleaned-name index for O(1) fuzzy lookups (built once at startup)
_CLEAN_NAME_INDEX: Dict[str, str] = {}
for _en_name, _zh_name in CHEMICAL_NAMES_ZH_EXPANDED.items():
    _clean = re.sub(r'[^a-z0-9]', '', _en_name)
    _CLEAN_NAME_INDEX[_clean] = _zh_name

# Define Models
class CASQuery(BaseModel):
    cas_numbers: List[str] = Field(..., max_length=100)

class GHSReport(BaseModel):
    """Single GHS classification report"""
    pictograms: List[Dict[str, Any]] = []
    hazard_statements: List[Dict[str, str]] = []
    precautionary_statements: List[Dict[str, str]] = []
    signal_word: Optional[str] = None
    signal_word_zh: Optional[str] = None
    source: Optional[str] = None
    report_count: Optional[str] = None

class ChemicalResult(BaseModel):
    cas_number: str
    cid: Optional[int] = None
    name_en: Optional[str] = None
    name_zh: Optional[str] = None
    # Primary (main) classification - first/most reported
    ghs_pictograms: List[Dict[str, Any]] = []
    hazard_statements: List[Dict[str, str]] = []
    precautionary_statements: List[Dict[str, str]] = []
    signal_word: Optional[str] = None
    signal_word_zh: Optional[str] = None
    # Other classification reports
    other_classifications: List[GHSReport] = []
    has_multiple_classifications: bool = False
    found: bool = False
    error: Optional[str] = None
    # True when the lookup failed because PubChem was temporarily
    # unavailable (timeout, 5xx, 429). Distinguishes "transient
    # upstream failure — please retry" from "CAS truly not in PubChem".
    # Frontend can display a different message for each case.
    upstream_error: bool = False
    # ── Provenance / trust signals (v1.8 M1) ──
    # Which primary-report source PubChem attributed the hazard data
    # to (e.g. "ECHA C&L Notifications Summary"). Pulled from the
    # top-ranked classification so users can gauge evidence strength
    # at a glance without opening the comparison table.
    primary_source: Optional[str] = None
    # How many notifier reports back the primary classification, if
    # PubChem exposed a count (e.g. "236 of 412 notifications from
    # companies ... with a classification ..."). Parsed as a plain
    # string because PubChem does not always express it as an int.
    primary_report_count: Optional[str] = None
    # ISO-8601 UTC timestamp when the GHS payload was fetched (or
    # last refreshed from PubChem). Helps the UI display "data as of
    # ..." so users can judge freshness.
    retrieved_at: Optional[str] = None
    # True when the GHS data was served from the 24hr TTL cache rather
    # than a fresh upstream fetch. The frontend can surface this as a
    # "cached" badge so users know to refresh if they need the very
    # latest data.
    cache_hit: bool = False
    # Enriched reference links combining built-in sources (PubChem /
    # ECHA / NIOSH) with any manually-added single-tenant links from
    # the pilot backend.
    reference_links: List[Dict[str, Any]] = []

# Maximum rows accepted by the export endpoints. Keeps request size
# bounded and prevents abuse of the export pipeline as a DoS vector.
MAX_EXPORT_ROWS = 500

class ExportRequest(BaseModel):
    results: List[Dict[str, Any]] = Field(..., max_length=MAX_EXPORT_ROWS)
    format: str = "xlsx"  # xlsx or csv


class WorkspaceDocumentPayload(BaseModel):
    payload: Any


class DictionaryMissQueryPayload(BaseModel):
    query: str
    query_kind: str = "name"
    endpoint: str = "frontend"
    context: Dict[str, Any] = Field(default_factory=dict)


class DictionaryManualEntryPayload(BaseModel):
    cas_number: str
    name_en: Optional[str] = None
    name_zh: Optional[str] = None
    notes: str = ""
    source: str = "manual"


class DictionaryAliasPayload(BaseModel):
    alias_text: str
    locale: str
    cas_number: str
    source: str = "manual"
    confidence: float = 1.0
    status: str = APPROVED_ALIAS_STATUS
    notes: str = ""


class DictionaryReferenceLinkPayload(BaseModel):
    cas_number: str
    label: str
    url: str
    link_type: str = "reference"
    source: str = "manual"
    priority: int = 50
    status: str = "active"
    cid: Optional[int] = None


# Characters that Excel / Google Sheets / LibreOffice Calc treat as the
# start of a formula. Values beginning with any of these can execute
# content when the exported file is opened elsewhere (CSV injection).
_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


def spreadsheet_safe(value: Any) -> str:
    """Neutralize spreadsheet formula injection in exported cells.

    Any string whose first character would be interpreted as a formula
    prefix is prefixed with an apostrophe, which spreadsheet apps treat
    as a text-only marker. `None` is coerced to an empty string.
    """
    if value is None:
        return ""
    text = str(value)
    if text and text[0] in _FORMULA_TRIGGER_CHARS:
        return "'" + text
    return text

# ─── PubChem resilience helper ──────────────────────────────
#
# Previously every PubChem call caught `Exception` and returned `{}`,
# so a 429 or 503 was indistinguishable from "no GHS data for this
# compound". For a chemical-safety tool that silent degradation is
# dangerous: users can read "no hazard" when really the upstream was
# briefly unavailable.
#
# `pubchem_get_json()` retries transient errors with exponential
# backoff and jitter, respects Retry-After on 429, and raises
# `PubChemError` when retries are exhausted so callers can decide
# whether to surface an upstream_error to the user.

class PubChemError(Exception):
    """Raised when PubChem is transiently unavailable after retries."""


_PUBCHEM_TRANSIENT_STATUS = {429, 500, 502, 503, 504}


async def pubchem_get_json(
    http_client: httpx.AsyncClient,
    url: str,
    *,
    timeout: float,
    retries: int = 2,
    max_delay: float = 4.0,
):
    """GET JSON from PubChem with retry/backoff.

    Returns
    -------
    (status_code, parsed_json_or_None)
        - (200, dict)         : success
        - (404, None)         : resource truly does not exist (no retry)
        - (4xx_other, None)   : treat as no-data (no retry)

    Raises
    ------
    PubChemError
        All retries exhausted on transient errors (timeout / 429 / 5xx /
        network). Callers should treat this as "upstream unavailable",
        NOT "no hazard data".
    """
    attempt = 0
    last_error = "unknown"
    while True:
        retry_after: Optional[str] = None
        transient = False
        # Bound the number of concurrent outbound PubChem requests so
        # a single burst of client traffic cannot balloon into a DoS
        # of PubChem (and get our IP rate-limited for everyone).
        async with _pubchem_semaphore:
            try:
                resp = await http_client.get(url, timeout=timeout)
            except httpx.TimeoutException as exc:
                last_error = f"{type(exc).__name__}: {exc}"
                transient = True
                _record_upstream_failure(
                    "timeout",
                    url,
                    attempt + 1,
                    detail=type(exc).__name__,
                )
                resp = None
            except httpx.TransportError as exc:
                last_error = f"{type(exc).__name__}: {exc}"
                transient = True
                _record_upstream_failure(
                    "transport_error",
                    url,
                    attempt + 1,
                    detail=type(exc).__name__,
                )
                resp = None
        if resp is not None:
            status = resp.status_code
            if status == 200:
                try:
                    return status, resp.json()
                except ValueError:
                    # 200 with non-JSON body — treat as no usable data
                    return status, None
            if status in _PUBCHEM_TRANSIENT_STATUS:
                transient = True
                if status == 429:
                    retry_after = resp.headers.get("Retry-After")
                    _record_upstream_failure(
                        "http_429",
                        url,
                        attempt + 1,
                        status_code=status,
                    )
                elif status >= 500:
                    _record_upstream_failure(
                        "http_5xx",
                        url,
                        attempt + 1,
                        status_code=status,
                    )
                last_error = f"HTTP {status}"
            else:
                # 4xx other than 429 (incl. 404) — definitive, no retry
                return status, None

        if not transient:
            # Shouldn't reach here, but guard against logic drift.
            raise PubChemError(f"{url}: {last_error}")

        attempt += 1
        if attempt > retries:
            _record_ops_counter("upstream.retry_exhausted")
            raise PubChemError(f"{url} failed after {attempt} attempts: {last_error}")

        # Exponential backoff with jitter; honour Retry-After when sensible.
        delay = min(max_delay, 0.3 * (2 ** (attempt - 1)))
        if retry_after:
            try:
                delay = max(delay, min(max_delay, float(retry_after)))
            except (TypeError, ValueError):
                pass
        delay += random.uniform(0, 0.15)
        await asyncio.sleep(delay)


# Helper functions
def normalize_cas(cas: str) -> str:
    """Normalize CAS number format"""
    cas = cas.strip()
    # Remove common prefixes
    cas = re.sub(r'^CAS[:\s-]*', '', cas, flags=re.IGNORECASE)
    # Keep only digits and hyphens
    cas = re.sub(r'[^\d-]', '', cas)
    
    # Try to fix common format issues
    if cas:
        parts = cas.split('-')
        if len(parts) == 3:
            # Remove leading zeros from first part
            parts[0] = parts[0].lstrip('0') or '0'
            # Remove leading zeros from last part (check digit should be single digit)
            parts[2] = parts[2].lstrip('0') or '0'
            # Ensure middle part has 2 digits
            if len(parts[1]) == 1:
                parts[1] = '0' + parts[1]
            cas = '-'.join(parts)
        elif len(parts) == 1 and len(cas) >= 5:
            # Try to parse CAS without hyphens (e.g., 6417-5 or 64175)
            digits = re.sub(r'[^0-9]', '', cas)
            if len(digits) >= 5:
                # CAS format: XXXXXX-XX-X
                check = digits[-1]
                middle = digits[-3:-1]
                first = digits[:-3].lstrip('0') or '0'
                cas = f"{first}-{middle}-{check}"
    
    return cas


def resolve_name_to_cas(query: str) -> Optional[str]:
    """Resolve English or Chinese chemical name to CAS number.
    Returns the first matching CAS number, or None.
    Priority: manual exact formal name → seed exact formal name →
              approved exact alias → compact exact match →
              word-boundary English → unique prefix English → unique prefix Chinese
    """
    q = query.strip()
    if not q:
        return None

    locale = infer_locale(q)

    manual_entry = pilot_store.get_manual_entry_by_name(q, locale, allow_compact=True)
    if manual_entry:
        return manual_entry["cas_number"]

    # Try exact Chinese name match (includes aliases like 酒精, 漂白水, etc.)
    if q in ZH_TO_CAS:
        return ZH_TO_CAS[q]

    # Try exact English name match (case-insensitive, includes aliases like "bleach", "dmso")
    q_lower = q.lower()
    if q_lower in EN_TO_CAS:
        return EN_TO_CAS[q_lower]

    exact_alias = pilot_store.get_alias_exact(q, locale)
    if exact_alias:
        return exact_alias["cas_number"]

    compact_match = _compact_exact_match(q, locale)
    if compact_match:
        return compact_match

    en_lookup = _combined_lookup_map("en")
    zh_lookup = _combined_lookup_map("zh")

    # Try English names containing the query as a word boundary match
    # e.g., "Methanol" matches "Methyl alcohol (Methanol)"
    en_contains = []
    for name, cas in en_lookup.items():
        # Check if query appears as a word (bounded by space, parens, or start/end)
        if re.search(r'(?:^|[\s(])' + re.escape(q_lower) + r'(?:$|[\s)])', name):
            en_contains.append(cas)
    if len(en_contains) == 1:
        return en_contains[0]

    # Try partial match (prefix) — English (case-insensitive)
    en_prefix = [cas for name, cas in en_lookup.items() if name.startswith(q_lower)]
    if len(en_prefix) == 1:
        return en_prefix[0]

    # Try partial match (prefix) — Chinese
    zh_prefix = [cas for name, cas in zh_lookup.items() if name.startswith(q)]
    if len(zh_prefix) == 1:
        return zh_prefix[0]

    return None


def _classification_signature(report: Dict[str, Any]) -> tuple:
    """Stable signature for deduplicating GHS classification reports.

    Two reports collapse into one only when ALL of the following match:
      - set of pictogram codes
      - signal word (e.g. Danger / Warning)
      - sorted set of H-statement codes
      - source string (e.g. ECHA C&L Notifications Summary)

    Previously dedup was keyed on pictogram set only, which dropped
    materially different classifications that shared the same icons
    but differed in hazard codes / signal word. That was a real data
    correctness issue for a chemical safety tool.
    """
    pic_codes = frozenset((p.get("code") or "") for p in report.get("pictograms", []))
    h_codes = tuple(sorted((h.get("code") or "") for h in report.get("hazard_statements", [])))
    p_codes = tuple(sorted((p.get("code") or "") for p in report.get("precautionary_statements", [])))
    signal = (report.get("signal_word") or "").strip()
    source = (report.get("source") or "").strip()
    return (pic_codes, signal, h_codes, p_codes, source)


def _report_rank_key(report: Dict[str, Any], source_index: int) -> tuple:
    """Deterministic rank key for choosing the primary classification.

    Lower tuple ranks first (ascending sort), so values that should
    rank earlier are negated.

    Priority, highest first:
      1. Larger ECHA report_count (stronger evidence base)
      2. Reports labelled as ECHA C&L Notifications (regulatory source)
      3. More hazard statements (more complete classification)
      4. Original PubChem source order as the stable tie-breaker
    """
    raw = report.get("report_count")
    try:
        count = int(raw) if raw else 0
    except (TypeError, ValueError):
        count = 0
    source = (report.get("source") or "").lower()
    echa_bonus = 1 if "echa" in source else 0
    hazard_count = len(report.get("hazard_statements") or [])
    return (-count, -echa_bonus, -hazard_count, source_index)


def extract_all_ghs_classifications(ghs_data: dict) -> List[Dict[str, Any]]:
    """
    Extract ALL GHS classification reports from PubChem data.
    Returns a list of classification reports, each containing pictograms, hazards, signal word, and source.
    The first report is typically the primary/most common classification.
    """
    reports = []
    
    try:
        sections = ghs_data.get("Record", {}).get("Section", [])
        for section in sections:
            if section.get("TOCHeading") == "Safety and Hazards":
                for subsection in section.get("Section", []):
                    if subsection.get("TOCHeading") == "Hazards Identification":
                        for subsubsection in subsection.get("Section", []):
                            if subsubsection.get("TOCHeading") == "GHS Classification":
                                # Parse each report entry
                                current_report = None
                                
                                for info in subsubsection.get("Information", []):
                                    info_name = info.get("Name", "")
                                    
                                    if info_name == "Pictogram(s)":
                                        # Start a new report when we encounter Pictogram(s)
                                        if current_report is not None:
                                            reports.append(current_report)
                                        
                                        current_report = {
                                            "pictograms": [],
                                            "hazard_statements": [],
                                            "precautionary_statements": [],
                                            "signal_word": None,
                                            "signal_word_zh": None,
                                            "source": None,
                                            "report_count": None
                                        }
                                        
                                        # Extract pictogram codes
                                        seen_codes = set()
                                        for markup in info.get("Value", {}).get("StringWithMarkup", []):
                                            for extra in markup.get("Markup", []):
                                                if extra.get("Type") == "Icon":
                                                    url = extra.get("URL", "")
                                                    match = re.search(r'(GHS\d{2})', url)
                                                    if match:
                                                        pic_code = match.group(1)
                                                        if pic_code in GHS_PICTOGRAMS and pic_code not in seen_codes:
                                                            seen_codes.add(pic_code)
                                                            current_report["pictograms"].append({
                                                                "code": pic_code,
                                                                **GHS_PICTOGRAMS[pic_code]
                                                            })
                                    
                                    elif info_name == "Signal" and current_report is not None:
                                        signal_translations = {"Danger": "危險", "Warning": "警告"}
                                        for markup in info.get("Value", {}).get("StringWithMarkup", []):
                                            signal = markup.get("String", "")
                                            if signal:
                                                current_report["signal_word"] = signal
                                                current_report["signal_word_zh"] = signal_translations.get(signal, signal)
                                                break
                                    
                                    elif info_name == "GHS Hazard Statements" and current_report is not None:
                                        seen_codes = set()
                                        for markup in info.get("Value", {}).get("StringWithMarkup", []):
                                            text = markup.get("String", "")
                                            h_match = re.search(r'(H\d{3})', text)
                                            if h_match:
                                                h_code = h_match.group(1)
                                                if h_code not in seen_codes:
                                                    seen_codes.add(h_code)
                                                    zh_text = H_CODE_TRANSLATIONS.get(h_code, "")
                                                    current_report["hazard_statements"].append({
                                                        "code": h_code,
                                                        "text_en": text,
                                                        "text_zh": zh_text if zh_text else text
                                                    })
                                    
                                    elif info_name == "Precautionary Statement Codes" and current_report is not None:
                                        # PubChem returns P-codes as a comma-separated
                                        # list in a single StringWithMarkup.String,
                                        # e.g. "P210, P233, P240, P303+P361+P353, and P501".
                                        # Combined codes like P301+P310 are kept intact.
                                        seen_p = set()
                                        for markup in info.get("Value", {}).get("StringWithMarkup", []):
                                            text = markup.get("String", "")
                                            # Extract every P-code token (single or combined)
                                            for p_match in re.finditer(r'(P\d{3}(?:\+P\d{3})*)', text):
                                                p_code = p_match.group(1)
                                                if p_code not in seen_p:
                                                    seen_p.add(p_code)
                                                    en_text = P_CODE_TEXTS_EN.get(p_code, "")
                                                    zh_text = P_CODE_TRANSLATIONS.get(p_code, "")
                                                    current_report["precautionary_statements"].append({
                                                        "code": p_code,
                                                        "text_en": en_text if en_text else p_code,
                                                        "text_zh": zh_text if zh_text else p_code,
                                                    })

                                    elif info_name == "ECHA C&L Notifications Summary" and current_report is not None:
                                        for markup in info.get("Value", {}).get("StringWithMarkup", []):
                                            text = markup.get("String", "")
                                            if text:
                                                current_report["source"] = text
                                                # Extract report count if available
                                                count_match = re.search(r'(\d+)\s*report', text.lower())
                                                if count_match:
                                                    current_report["report_count"] = count_match.group(1)
                                                break
                                
                                # Don't forget the last report
                                if current_report is not None:
                                    reports.append(current_report)
    
    except Exception as e:
        logger.error(f"Error extracting GHS classifications: {e}")
    
    # Filter out empty reports
    reports = [r for r in reports if r.get("pictograms")]
    
    return reports


def get_chinese_name_from_cas(cas_number: str) -> Optional[str]:
    """Get Chinese name directly from CAS number (most accurate method)"""
    if not cas_number:
        return None
    cas_normalized = cas_number.strip()

    manual = pilot_store.get_manual_entry_by_cas(cas_normalized)
    if manual and manual.get("name_zh"):
        return manual["name_zh"]
    
    # Direct CAS lookup - highest priority
    if cas_normalized in CAS_TO_ZH:
        return CAS_TO_ZH[cas_normalized]
    
    return None

def get_english_name_from_cas(cas_number: str) -> Optional[str]:
    """Get English name from CAS number via local dictionary"""
    if not cas_number:
        return None
    cas_normalized = cas_number.strip()

    manual = pilot_store.get_manual_entry_by_cas(cas_normalized)
    if manual and manual.get("name_en"):
        return manual["name_en"]
    
    # Direct lookup from CAS_TO_EN dictionary
    if cas_normalized in CAS_TO_EN:
        return CAS_TO_EN[cas_normalized]
    
    return None

def get_chinese_name_from_dict(name_en: str) -> Optional[str]:
    """Get Chinese name from English name dictionary (optimized with pre-built index)."""
    if not name_en:
        return None
    name_lower = name_en.lower().strip()

    manual = pilot_store.get_manual_entry_by_name(name_en, "en", allow_compact=True)
    if manual and manual.get("name_zh"):
        return manual["name_zh"]

    # O(1) exact match in expanded dictionary (1861 entries)
    if name_lower in CHEMICAL_NAMES_ZH_EXPANDED:
        return CHEMICAL_NAMES_ZH_EXPANDED[name_lower]

    # O(1) cleaned-name match via pre-built index
    name_clean = re.sub(r'[^a-z0-9]', '', name_lower)
    if name_clean in _CLEAN_NAME_INDEX:
        return _CLEAN_NAME_INDEX[name_clean]

    return None

async def _try_cid_by_name(cas_number: str, http_client: httpx.AsyncClient) -> Optional[int]:
    """CID lookup Method 1: Search by CAS number as compound name.

    Transient failures propagate as PubChemError so get_cid_from_cas()
    can distinguish "all methods failed due to upstream outage" from
    "PubChem returned 404 on every method (truly unknown CAS)".
    """
    url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{cas_number}/cids/JSON"
    status, data = await pubchem_get_json(http_client, url, timeout=15.0)
    if status == 200 and data:
        cids = data.get("IdentifierList", {}).get("CID", [])
        if cids:
            return cids[0]
    return None

async def _try_cid_by_xref(cas_number: str, http_client: httpx.AsyncClient) -> Optional[int]:
    """CID lookup Method 2: Search via xref/rn endpoint."""
    url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/xref/rn/{cas_number}/cids/JSON"
    status, data = await pubchem_get_json(http_client, url, timeout=15.0)
    if status == 200 and data:
        cids = data.get("IdentifierList", {}).get("CID", [])
        if cids:
            return cids[0]
    return None

async def _try_cid_by_substance(cas_number: str, http_client: httpx.AsyncClient) -> Optional[int]:
    """CID lookup Method 3: Search via substance xref."""
    url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/substance/xref/rn/{cas_number}/cids/JSON"
    status, data = await pubchem_get_json(http_client, url, timeout=15.0)
    if status == 200 and data:
        info = data.get("InformationList", {}).get("Information", [])
        if info and info[0].get("CID"):
            cid_field = info[0]["CID"]
            return cid_field[0] if isinstance(cid_field, list) else cid_field
    return None

async def get_cid_from_cas(cas_number: str, http_client: httpx.AsyncClient) -> Optional[int]:
    """Get PubChem CID from CAS number — runs methods 1-3 concurrently, with cache.

    Policy for a safety-critical tool:

      1. If any method returns an int, use it.
      2. Otherwise, if ANY method (primary or alt-CAS fallback) raised
         `PubChemError` — i.e. at least one endpoint transiently failed
         — the "not found" conclusion is not trustworthy, so we raise
         `PubChemError` and let the caller surface `upstream_error=True`.
      3. Only when every method returned a clean "no match" do we
         return `None` to mean "truly not in PubChem".

    Previously the function only raised when ALL three primary methods
    were transient failures, which let the mixed case through
    (e.g. one 404 + two 503) and could present a transient outage as a
    confirmed absence. That is unsafe for a GHS lookup tool.
    """
    # Check cache first
    if cas_number in cid_cache:
        return cid_cache[cas_number]

    # Run methods 1-3 concurrently; return_exceptions so one transient
    # failure does not abort the others.
    results = await asyncio.gather(
        _try_cid_by_name(cas_number, http_client),
        _try_cid_by_xref(cas_number, http_client),
        _try_cid_by_substance(cas_number, http_client),
        return_exceptions=True,
    )
    for result in results:
        if isinstance(result, int):
            cid_cache[cas_number] = result
            return result

    primary_had_transient = any(isinstance(r, PubChemError) for r in results)

    # Method 4 (fallback): alternate CAS format (strip leading zeros).
    # We attempt it whether or not the primary results had transient
    # failures — it might still succeed and give us a definitive CID.
    fallback_had_transient = False
    cas_alt = re.sub(r'^0+', '', cas_number.split('-')[0]) + '-' + '-'.join(cas_number.split('-')[1:])
    if cas_alt != cas_number:
        try:
            cid = await _try_cid_by_name(cas_alt, http_client)
        except PubChemError:
            cid = None
            fallback_had_transient = True
        if cid:
            cid_cache[cas_number] = cid
            return cid

    # No CID found anywhere. If ANY attempt (primary or fallback) was
    # a transient failure, refuse to commit to "not found".
    if primary_had_transient or fallback_had_transient:
        raise PubChemError(
            f"CID lookup for {cas_number} had partial upstream failures; "
            "cannot confirm not-found."
        )

    logger.info(f"PubChem has no CID for CAS number: {cas_number}")
    return None

async def get_compound_name(
    cid: int,
    http_client: httpx.AsyncClient,
    known_zh: Optional[str] = None,
    cas_number: Optional[str] = None,
) -> tuple:
    """Get compound name in English and Chinese with multiple fallbacks.

    Name resolution is explicitly best-effort: a transient PubChem
    failure here is not fatal (we can still fall back to the local
    dictionary and to the RecordTitle inside the GHS payload). So we
    catch PubChemError per endpoint instead of propagating.
    If known_zh is provided, skip expensive Chinese name lookups."""
    name_en = None
    name_zh = known_zh
    all_synonyms = []

    # Method 1: Get from property endpoint
    try:
        url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/property/IUPACName,Title/JSON"
        status, data = await pubchem_get_json(http_client, url, timeout=15.0)
        if status == 200 and data:
            props = data.get("PropertyTable", {}).get("Properties", [{}])[0]
            name_en = props.get("Title") or props.get("IUPACName")
    except PubChemError as e:
        logger.debug(f"Property endpoint transient failure for CID {cid}: {e}")

    # Method 2: Get from synonyms endpoint
    try:
        syn_url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/synonyms/JSON"
        status, syn_data = await pubchem_get_json(http_client, syn_url, timeout=15.0)
        if status == 200 and syn_data:
            all_synonyms = syn_data.get("InformationList", {}).get("Information", [{}])[0].get("Synonym", [])

            # If no name_en yet, use first synonym
            if not name_en and all_synonyms:
                name_en = all_synonyms[0]

            # Look for Chinese characters in synonyms
            for syn in all_synonyms:
                if any('\u4e00' <= char <= '\u9fff' for char in syn):
                    name_zh = syn
                    break
    except PubChemError as e:
        logger.debug(f"Synonyms endpoint transient failure for CID {cid}: {e}")

    # Method 3: Try description endpoint for name
    if not name_en:
        try:
            desc_url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/description/JSON"
            status, desc_data = await pubchem_get_json(http_client, desc_url, timeout=15.0)
            if status == 200 and desc_data:
                info_list = desc_data.get("InformationList", {}).get("Information", [])
                if info_list:
                    name_en = info_list[0].get("Title")
        except PubChemError as e:
            logger.debug(f"Description endpoint transient failure for CID {cid}: {e}")

    # If no Chinese name yet, try local dictionary lookups
    if not name_zh and name_en:
        name_zh = get_chinese_name_from_dict(name_en)
    if not name_zh:
        for syn in all_synonyms[:15]:
            zh = get_chinese_name_from_dict(syn)
            if zh:
                name_zh = zh
                break

    if cas_number and all_synonyms:
        excluded = {
            normalize_compact_text(known_zh, locale="zh"),
            normalize_compact_text(name_en, locale="en"),
            normalize_compact_text(name_zh, locale="zh"),
        }
        candidate_synonyms = []
        for synonym in all_synonyms[:25]:
            locale = infer_locale(synonym)
            compact = normalize_compact_text(synonym, locale=locale)
            if not compact or compact in excluded:
                continue
            candidate_synonyms.append(synonym)
        pilot_store.capture_alias_candidates(cas_number, candidate_synonyms)

    return name_en, name_zh

def extract_record_title(ghs_data: dict) -> str:
    """Extract RecordTitle from GHS data as fallback name"""
    try:
        return ghs_data.get("Record", {}).get("RecordTitle", "")
    except:
        return ""

def extract_iupac_name(ghs_data: dict) -> str:
    """Extract IUPAC name from GHS data as another fallback"""
    try:
        sections = ghs_data.get("Record", {}).get("Section", [])
        for section in sections:
            if section.get("TOCHeading") == "Names and Identifiers":
                for subsection in section.get("Section", []):
                    if subsection.get("TOCHeading") == "Computed Descriptors":
                        for subsubsection in subsection.get("Section", []):
                            if subsubsection.get("TOCHeading") == "IUPAC Name":
                                info = subsubsection.get("Information", [{}])[0]
                                return info.get("Value", {}).get("StringWithMarkup", [{}])[0].get("String", "")
    except:
        pass
    return ""

async def get_ghs_classification(cid: int, http_client: httpx.AsyncClient) -> tuple:
    """Get GHS classification from PubChem (with 24hr cache).

    Unlike the name endpoints, transient failures here are critical:
    returning `{}` would cause the rest of search_chemical() to emit
    a found=True result with empty hazard data, which is dangerous
    for a safety tool. So we let PubChemError propagate to the caller
    and only return `({}, ...)` when PubChem definitively says there's
    no GHS section for this CID (e.g. 404).

    Returns
    -------
    (ghs_data, cache_hit, retrieved_at)
        - ghs_data     : dict (empty if no GHS section exists)
        - cache_hit    : bool (True if served from ghs_cache)
        - retrieved_at : ISO-8601 UTC timestamp string recording when
                         we fetched or last refreshed this entry

    The cache value is now `(data, retrieved_at)` so provenance can be
    surfaced to the user ("data fetched at X / served from cache").
    """
    cached = ghs_cache.get(cid)
    if cached is not None:
        data, retrieved_at = cached
        _observe_ghs_cache_hit(cid, retrieved_at)
        return data, True, retrieved_at

    _record_ops_counter("cache.ghs.miss")

    url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON"
    status, data = await pubchem_get_json(http_client, url, timeout=30.0)
    now = datetime.now(timezone.utc).isoformat()
    if status == 200 and data:
        ghs_cache[cid] = (data, now)
        return data, False, now
    # 404 / empty: don't cache a fresh miss (keeps retry-able) but still
    # report the current timestamp so the caller can annotate the result.
    return {}, False, now

async def search_chemical(cas_number: str, http_client: httpx.AsyncClient) -> ChemicalResult:
    """Search for a chemical by CAS number"""
    normalized_cas = normalize_cas(cas_number)

    if not normalized_cas:
        return ChemicalResult(
            cas_number=cas_number,
            found=False,
            error="無效的 CAS 號碼格式（正確格式如：64-17-5）"
        )

    # Validate CAS number format (should be like XX-XX-X or XXXXX-XX-X)
    cas_pattern = re.match(r'^(\d{2,7})-(\d{2})-(\d)$', normalized_cas)
    if not cas_pattern:
        return ChemicalResult(
            cas_number=cas_number,
            found=False,
            error=f"CAS 號碼格式不正確：{normalized_cas}（正確格式如：64-17-5）"
        )

    # ===== NEW: Try to get Chinese and English name from CAS dictionary FIRST (most accurate) =====
    name_zh_from_cas = get_chinese_name_from_cas(normalized_cas)
    name_en_from_cas = get_english_name_from_cas(normalized_cas)
    if name_zh_from_cas:
        logger.info(f"Found Chinese name from CAS dictionary for {normalized_cas}: {name_zh_from_cas}")
    if name_en_from_cas:
        logger.info(f"Found English name from CAS dictionary for {normalized_cas}: {name_en_from_cas}")

    # Get CID - try multiple methods. A transient upstream failure must
    # NOT collapse into "not found" — surface it as upstream_error so
    # the frontend can tell the user to retry rather than assume the
    # chemical has no hazard data.
    try:
        cid = await get_cid_from_cas(normalized_cas, http_client)
    except PubChemError as e:
        logger.warning(f"PubChem unavailable during CID lookup for {normalized_cas}: {e}")
        return ChemicalResult(
            cas_number=cas_number,
            name_en=name_en_from_cas,
            name_zh=name_zh_from_cas,
            found=False,
            upstream_error=True,
            error="PubChem 暫時無法回應，請稍後再試 (CID lookup failed)"
        )
    if not cid:
        # Even if PubChem doesn't have CID, we might have local data
        if name_zh_from_cas or name_en_from_cas:
            return ChemicalResult(
                cas_number=cas_number,
                cid=None,
                name_en=name_en_from_cas,  # Provide English name from local dictionary
                name_zh=name_zh_from_cas,
                ghs_pictograms=[],
                hazard_statements=[],
                precautionary_statements=[],
                signal_word=None,
                signal_word_zh=None,
                found=True,
                reference_links=_build_reference_links(normalized_cas, None, name_en_from_cas),
                error="PubChem 無 GHS 資料，僅提供本地字典名稱"
            )
        # Provide more helpful error message
        return ChemicalResult(
            cas_number=cas_number,
            found=False,
            error=f"在 PubChem 資料庫中找不到 CAS {normalized_cas}，請確認號碼是否正確"
        )
    
    # Get compound name and GHS data concurrently
    # Pass known Chinese name to skip redundant dictionary lookups.
    # Critical: a transient failure of the GHS endpoint must not be
    # silently rendered as "no hazards". `get_ghs_classification()`
    # raises PubChemError for transient cases; catch it here and
    # return an upstream_error result so the UI can distinguish
    # "retry later" from "this chemical has no GHS data".
    name_task = get_compound_name(
        cid,
        http_client,
        known_zh=name_zh_from_cas,
        cas_number=normalized_cas,
    )
    ghs_task = get_ghs_classification(cid, http_client)

    try:
        name_result, ghs_result = await asyncio.gather(name_task, ghs_task)
        name_en, name_zh = name_result
        ghs_data, cache_hit, retrieved_at = ghs_result
    except PubChemError as e:
        logger.warning(f"PubChem unavailable during GHS lookup for CID {cid}: {e}")
        return ChemicalResult(
            cas_number=cas_number,
            cid=cid,
            name_en=name_en_from_cas,
            name_zh=name_zh_from_cas,
            found=False,
            upstream_error=True,
            error="PubChem 暫時無法回應，請稍後再試 (GHS classification fetch failed)"
        )
    
    # Extract ALL GHS classification reports
    all_classifications = extract_all_ghs_classifications(ghs_data)
    
    # Separate primary (first) classification from others
    primary_pictograms = []
    primary_hazards = []
    primary_precautions = []
    primary_signal = None
    primary_signal_zh = None
    primary_source: Optional[str] = None
    primary_report_count: Optional[str] = None
    other_classifications = []

    if all_classifications:
        # Rank all reports deterministically; the top-ranked is primary.
        indexed = list(enumerate(all_classifications))
        indexed.sort(key=lambda ix: _report_rank_key(ix[1], ix[0]))

        _, primary = indexed[0]
        primary_pictograms = primary.get("pictograms", [])
        primary_hazards = primary.get("hazard_statements", [])
        primary_precautions = primary.get("precautionary_statements", [])
        primary_signal = primary.get("signal_word")
        primary_signal_zh = primary.get("signal_word_zh")
        primary_source = primary.get("source")
        primary_report_count = primary.get("report_count")

        # Dedup using the full signature (pictograms + signal + h-codes + source).
        seen_signatures = {_classification_signature(primary)}
        for _, report in indexed[1:]:
            if not report.get("pictograms"):
                continue
            sig = _classification_signature(report)
            if sig in seen_signatures:
                continue
            seen_signatures.add(sig)
            other_classifications.append(GHSReport(
                pictograms=report.get("pictograms", []),
                hazard_statements=report.get("hazard_statements", []),
                precautionary_statements=report.get("precautionary_statements", []),
                signal_word=report.get("signal_word"),
                signal_word_zh=report.get("signal_word_zh"),
                source=report.get("source"),
                report_count=report.get("report_count"),
            ))
    
    # Use RecordTitle as fallback for name_en if not found
    if not name_en:
        name_en = extract_record_title(ghs_data)
    
    # Try IUPAC name as another fallback
    if not name_en:
        name_en = extract_iupac_name(ghs_data)
    
    # If still no name, use CAS number as name
    if not name_en:
        name_en = f"CID-{cid}"
        logger.warning(f"No name found for CAS {cas_number}, using CID as fallback")
    
    # ===== OPTIMIZED Chinese name lookup order =====
    # Priority 1: CAS dictionary (most accurate - from user's Excel)
    if name_zh_from_cas:
        name_zh = name_zh_from_cas
    # Priority 2: PubChem Chinese synonyms (already fetched)
    elif name_zh:
        pass  # Keep PubChem result
    # Priority 3: English name dictionary lookup
    else:
        name_zh = get_chinese_name_from_dict(name_en)
    
    return ChemicalResult(
        cas_number=cas_number,
        cid=cid,
        name_en=name_en,
        name_zh=name_zh,
        ghs_pictograms=primary_pictograms,
        hazard_statements=primary_hazards,
        precautionary_statements=primary_precautions,
        signal_word=primary_signal,
        signal_word_zh=primary_signal_zh,
        other_classifications=other_classifications,
        has_multiple_classifications=len(other_classifications) > 0,
        found=True,
        primary_source=primary_source,
        primary_report_count=primary_report_count,
        retrieved_at=retrieved_at,
        cache_hit=cache_hit,
        reference_links=_build_reference_links(normalized_cas, cid, name_en),
    )

# API Routes
@api_router.get("/")
async def root():
    return {"message": "GHS Label Quick Search API"}

@api_router.get("/health")
async def health_check():
    """Health check endpoint for monitoring and load balancers."""
    return {
        "status": "healthy",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "version": APP_VERSION,
    }


@api_router.get("/ops/report")
async def ops_report(request: Request):
    _require_admin(request)
    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "version": APP_VERSION,
        "staleThresholdHours": OPS_STALE_THRESHOLD_HOURS,
        "counters": dict(sorted(ops_counters.items())),
        "cache": {
            "cidEntries": len(cid_cache),
            "ghsEntries": len(ghs_cache),
        },
        "recentEvents": list(ops_recent_events),
        "dictionary": pilot_store.get_dictionary_summary(limit=10),
    }


@api_router.get("/dictionary/report")
async def dictionary_report(request: Request):
    _require_admin(request)
    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "version": APP_VERSION,
        "dictionary": pilot_store.get_dictionary_summary(limit=50),
    }


@api_router.get("/dictionary/manual-entries")
async def dictionary_manual_entries(request: Request):
    _require_admin(request)
    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "items": pilot_store.list_manual_entries(),
    }


@api_router.post("/dictionary/manual-entries")
async def upsert_dictionary_manual_entry(request: Request, payload: DictionaryManualEntryPayload):
    _require_admin(request)
    record = pilot_store.upsert_dictionary_entry(
        payload.cas_number,
        name_en=payload.name_en,
        name_zh=payload.name_zh,
        notes=payload.notes,
        source=payload.source,
    )
    return {"ok": True, "record": record}


@api_router.get("/dictionary/aliases")
async def dictionary_aliases(
    request: Request,
    status: Optional[str] = None,
    locale: Optional[str] = None,
    cas_number: Optional[str] = None,
):
    _require_admin(request)
    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "items": pilot_store.list_aliases(
            status=status,
            locale=locale,
            cas_number=cas_number,
        ),
    }


@api_router.post("/dictionary/aliases")
async def upsert_dictionary_alias(request: Request, payload: DictionaryAliasPayload):
    _require_admin(request)
    record = pilot_store.upsert_alias(
        payload.alias_text,
        payload.locale,
        payload.cas_number,
        source=payload.source,
        confidence=payload.confidence,
        status=payload.status,
        notes=payload.notes,
    )
    return {"ok": True, "record": record}


@api_router.get("/dictionary/reference-links")
async def dictionary_reference_links(
    request: Request,
    cas_number: Optional[str] = None,
    include_inactive: bool = False,
):
    _require_admin(request)
    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "items": pilot_store.list_reference_links(
            cas_number,
            include_inactive=include_inactive,
        ),
    }


@api_router.post("/dictionary/reference-links")
async def upsert_dictionary_reference_link(
    request: Request,
    payload: DictionaryReferenceLinkPayload,
):
    _require_admin(request)
    record = pilot_store.upsert_reference_link(
        payload.cas_number,
        label=payload.label,
        url=payload.url,
        link_type=payload.link_type,
        source=payload.source,
        priority=payload.priority,
        status=payload.status,
        cid=payload.cid,
    )
    return {"ok": True, "record": record}


@api_router.post("/dictionary/miss-query")
async def dictionary_miss_query(payload: DictionaryMissQueryPayload):
    if not CAPTURE_DICTIONARY_MISSES:
        return {
            "ok": False,
            "skipped": True,
            "reason": "dictionary miss capture is disabled",
        }

    return {
        "ok": True,
        "record": pilot_store.record_miss_query(
            payload.query,
            payload.query_kind,
            payload.endpoint,
            context=payload.context,
        ),
    }


@api_router.get("/workspace/{doc_type}")
async def get_workspace_document(request: Request, doc_type: str):
    _require_admin(request)
    doc_type = _ensure_workspace_doc_type(doc_type)
    document = pilot_store.get_document(doc_type)
    if document is None:
        return {
            "docType": doc_type,
            "payload": None,
            "updatedAt": None,
        }
    return {
        "docType": doc_type,
        "payload": document["payload"],
        "updatedAt": document["updatedAt"],
    }


@api_router.put("/workspace/{doc_type}")
async def put_workspace_document(
    request: Request,
    doc_type: str,
    payload: WorkspaceDocumentPayload,
):
    _require_admin(request)
    doc_type = _ensure_workspace_doc_type(doc_type)
    document = pilot_store.put_document(doc_type, payload.payload)
    return {
        "docType": doc_type,
        "payload": document["payload"],
        "updatedAt": document["updatedAt"],
    }

@api_router.post("/search", response_model=List[ChemicalResult])
@limiter.limit("10/minute")
async def search_chemicals(request: Request, query: CASQuery):
    """Search for chemicals by CAS numbers"""
    results = []
    http_client = shared_http_client
    # Process in batches to avoid overwhelming the API
    batch_size = 5
    for i in range(0, len(query.cas_numbers), batch_size):
        batch = query.cas_numbers[i:i+batch_size]
        tasks = [search_chemical(cas, http_client) for cas in batch]
        batch_results = await asyncio.gather(*tasks)
        results.extend(batch_results)
        # Add small delay between batches
        if i + batch_size < len(query.cas_numbers):
            await asyncio.sleep(0.5)
    for raw_query, result in zip(query.cas_numbers, results):
        if result.found or result.upstream_error:
            continue
        _record_dictionary_miss(
            raw_query,
            "cas",
            "search_batch",
            context={"normalizedCas": normalize_cas(raw_query)},
        )
    return results

@api_router.get("/search-by-name/{query}")
@limiter.limit("60/minute")
async def search_by_name(request: Request, query: str):
    """Search for chemicals by English or Chinese name (including aliases/common names).
    Returns list of matching {cas_number, name_en, name_zh, alias} (max 20).
    Used for autocomplete / name lookup before full GHS search.
    The `alias` field is set when matched via a common name (e.g., "酒精" → 乙醇).
    """
    q = query.strip()
    if not q or len(q) < 2:
        return {"results": [], "query": q}

    q_lower = q.lower()
    matches = []
    seen_rows = set()

    def append_match(cas: str, *, alias: Optional[str] = None) -> None:
        key = (cas, alias or "")
        if key in seen_rows:
            return
        seen_rows.add(key)
        matches.append({
            "cas_number": cas,
            "name_en": get_english_name_from_cas(cas) or "",
            "name_zh": get_chinese_name_from_cas(cas) or "",
            "alias": alias,
            "reference_links": _build_reference_links(cas, None, get_english_name_from_cas(cas)),
        })

    # Chinese name matches (substring) — includes seed aliases merged into ZH_TO_CAS
    for name, cas in ZH_TO_CAS.items():
        if q in name:
            append_match(cas, alias=name if name in ALIASES_ZH else None)

    # English name matches (case-insensitive substring) — includes seed aliases merged into EN_TO_CAS
    for name, cas in EN_TO_CAS.items():
        if q_lower in name:
            append_match(cas, alias=name if name in ALIASES_EN else None)

    for name, cas in _manual_name_pairs("zh"):
        if q in name:
            append_match(cas)

    for name, cas in _manual_name_pairs("en"):
        if q_lower in name:
            append_match(cas)

    for alias_text, cas in _approved_alias_pairs("zh"):
        if q in alias_text:
            append_match(cas, alias=alias_text)

    for alias_text, cas in _approved_alias_pairs("en"):
        if q_lower in alias_text:
            append_match(cas, alias=alias_text)

    # Sort: exact match first, then alias matches, then by name length
    matches.sort(key=lambda m: (
        0 if m["name_en"].lower() == q_lower or m["name_zh"] == q else
        0 if m.get("alias") and (m["alias"] == q or m["alias"].lower() == q_lower) else 1,
        len(m["name_en"])
    ))

    results = matches[:20]
    if len(results) == 0:
        _record_dictionary_miss(q, "autocomplete", "search_by_name")

    return {"results": results, "query": q}


async def _search_single_query(query: str) -> ChemicalResult:
    query = (query or "").strip()
    if not query:
        raise HTTPException(status_code=400, detail="Query is required.")

    # Check if it looks like a CAS number (digits and hyphens only)
    if re.match(r'^[\d-]+$', query):
        result = await search_chemical(query, shared_http_client)
        if not result.found and not result.upstream_error:
            _record_dictionary_miss(
                query,
                "cas",
                "search_single",
                context={"normalizedCas": normalize_cas(query)},
            )
        return result

    # Try to resolve name to CAS
    resolved_cas = resolve_name_to_cas(query)
    if resolved_cas:
        return await search_chemical(resolved_cas, shared_http_client)

    # Not found by name
    _record_dictionary_miss(query, "name", "search_single")
    return ChemicalResult(
        cas_number=query,
        found=False,
        error=f"No chemical found for name: {query}"
    )


@api_router.get("/search-single", response_model=ChemicalResult)
@limiter.limit("30/minute")
async def search_single_chemical_query(request: Request, q: str):
    return await _search_single_query(q)


@api_router.get("/search/{cas_number}", response_model=ChemicalResult)
@limiter.limit("30/minute")
async def search_single_chemical(request: Request, cas_number: str):
    """Search by CAS number or chemical name.
    Auto-detects whether input is a CAS number or name."""
    return await _search_single_query(cas_number)


@api_router.post("/export/xlsx")
@limiter.limit("10/minute")
async def export_xlsx(request: Request, payload: ExportRequest):
    """Export results to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "GHS查詢結果"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["CAS No.", "英文名稱", "中文名稱", "GHS標示", "警示語", "危害說明", "預防措施"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data — every cell value is routed through spreadsheet_safe() to
    # neutralize values that start with a formula prefix (e.g.
    # `=HYPERLINK(...)`, `+cmd`, `@SUM(...)`), preventing CSV/XLSX
    # injection when the file is opened in Excel / Sheets / Calc.
    for row, result in enumerate(payload.results, 2):
        ws.cell(row=row, column=1, value=spreadsheet_safe(result.get("cas_number", ""))).border = thin_border
        ws.cell(row=row, column=2, value=spreadsheet_safe(result.get("name_en", ""))).border = thin_border
        ws.cell(row=row, column=3, value=spreadsheet_safe(result.get("name_zh", ""))).border = thin_border

        # GHS pictograms
        pictograms = result.get("ghs_pictograms", [])
        ghs_text = ", ".join([f"{p.get('code', '')} ({p.get('name_zh', '')})" for p in pictograms]) if pictograms else "無"
        ws.cell(row=row, column=4, value=spreadsheet_safe(ghs_text)).border = thin_border

        # Signal word
        signal = result.get("signal_word_zh") or result.get("signal_word") or "-"
        ws.cell(row=row, column=5, value=spreadsheet_safe(signal)).border = thin_border

        # Hazard statements
        statements = result.get("hazard_statements", [])
        hazard_text = "\n".join([f"{s.get('code', '')}: {s.get('text_zh', '')}" for s in statements]) if statements else "無危害說明"
        hazard_cell = ws.cell(row=row, column=6, value=spreadsheet_safe(hazard_text))
        hazard_cell.border = thin_border
        hazard_cell.alignment = Alignment(wrap_text=True)

        # Precautionary statements (v1.8 M0)
        precautions = result.get("precautionary_statements", [])
        precaution_text = (
            "\n".join([f"{p.get('code', '')}: {p.get('text_zh', '')}" for p in precautions])
            if precautions
            else "無預防措施說明"
        )
        prec_cell = ws.cell(row=row, column=7, value=spreadsheet_safe(precaution_text))
        prec_cell.border = thin_border
        prec_cell.alignment = Alignment(wrap_text=True)

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 55
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ghs_results.xlsx"}
    )

@api_router.post("/export/csv")
@limiter.limit("10/minute")
async def export_csv(request: Request, payload: ExportRequest):
    """Export results to CSV file"""
    from io import StringIO
    string_output = StringIO()
    writer = csv.writer(string_output)
    
    # Headers
    writer.writerow(["CAS No.", "英文名稱", "中文名稱", "GHS標示", "警示語", "危害說明", "預防措施"])

    # Data
    for result in payload.results:
        pictograms = result.get("ghs_pictograms", [])
        ghs_text = ", ".join([f"{p.get('code', '')} ({p.get('name_zh', '')})" for p in pictograms]) if pictograms else "無"

        signal = result.get("signal_word_zh") or result.get("signal_word") or "-"

        statements = result.get("hazard_statements", [])
        hazard_text = "; ".join([f"{s.get('code', '')}: {s.get('text_zh', '')}" for s in statements]) if statements else "無危害說明"

        precautions = result.get("precautionary_statements", [])
        precaution_text = (
            "; ".join([f"{p.get('code', '')}: {p.get('text_zh', '')}" for p in precautions])
            if precautions
            else "無預防措施說明"
        )

        writer.writerow([
            spreadsheet_safe(result.get("cas_number", "")),
            spreadsheet_safe(result.get("name_en", "")),
            spreadsheet_safe(result.get("name_zh", "")),
            spreadsheet_safe(ghs_text),
            spreadsheet_safe(signal),
            spreadsheet_safe(hazard_text),
            spreadsheet_safe(precaution_text),
        ])
    
    # Convert to bytes with BOM
    csv_content = string_output.getvalue()
    output = BytesIO()
    output.write(b'\xef\xbb\xbf')  # BOM
    output.write(csv_content.encode('utf-8'))
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=ghs_results.csv"}
    )

@api_router.get("/ghs-pictograms")
async def get_ghs_pictograms():
    """Get all GHS pictogram information"""
    return GHS_PICTOGRAMS

# Include the router in the main app
app.include_router(api_router)

# CORS configuration.
#
# Defaults are intentionally strict:
#   - `allow_credentials=False` because the API does not use cookies,
#     `Authorization` headers, or any other credentialed browser flow.
#     Keeping it False means that even if an origin is mis-allowed, the
#     browser will refuse to attach credentials — narrowing blast radius.
#   - `allow_origins` falls back to the production frontend only.
#     Wildcard (`*`) is explicitly rejected here because it is unsafe
#     with credentials (and confusing if credentials are later enabled).
#
# Local development should set `CORS_ORIGINS=http://localhost:5173`
# (or a comma-separated list) in `.env` / Docker compose.
_raw_cors = os.environ.get("CORS_ORIGINS", "https://ghs-frontend.zeabur.app")
_cors_origins = [o.strip() for o in _raw_cors.split(",") if o.strip()]
if "*" in _cors_origins:
    logger.warning(
        "CORS_ORIGINS=* is unsafe for a public API; falling back to the "
        "production frontend. Set explicit origins for multi-origin deploys."
    )
    _cors_origins = ["https://ghs-frontend.zeabur.app"]

app.add_middleware(
    CORSMiddleware,
    allow_credentials=False,
    allow_origins=_cors_origins,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["Content-Type", "Accept"],
)


# ─── Security response headers ──────────────────────────────
#
# Add a small set of defence-in-depth headers to every API response.
# These are cheap, widely supported, and make common attacks harder:
#
#   - `X-Content-Type-Options: nosniff`
#     Browsers must not sniff and "guess" a different content-type
#     than what we send. Prevents a JSON payload being rendered as
#     HTML/JS if a Content-Type header is ever malformed.
#
#   - `Referrer-Policy: strict-origin-when-cross-origin`
#     When the frontend navigates away or issues cross-origin
#     requests, don't leak the full URL (including query strings,
#     which could contain a chemical name the user typed).
#
#   - `Permissions-Policy: camera=(), microphone=(), geolocation=()...`
#     Explicitly disable browser APIs this app never uses. Reduces
#     attack surface if an XSS were ever to land.
#
#   - `Content-Security-Policy: default-src 'none'; frame-ancestors 'none'`
#     The API returns JSON (and a binary xlsx/csv stream); it never
#     renders HTML. `default-src 'none'` means a browser that
#     somehow interpreted a response as HTML would execute no
#     scripts, load no images, etc. `frame-ancestors 'none'`
#     prevents API responses from being framed at all.
#
# The frontend is a static site and carries its own CSP via a meta
# tag in index.html (added alongside this change).
@app.middleware("http")
async def security_headers_middleware(request, call_next):
    response = await call_next(request)
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault(
        "Permissions-Policy",
        "camera=(), microphone=(), geolocation=(), payment=(), usb=()",
    )
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'none'; frame-ancestors 'none'",
    )
    return response

